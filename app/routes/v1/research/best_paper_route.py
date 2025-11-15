import re
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from flask import request, jsonify, current_app, abort, send_file
import json
import os
import uuid
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.models.User import User
from app.routes.v1.research import research_bp
from app.models.Cycle import BestPaperVerifiers, BestPaper, Author, Category, PaperCategory, Cycle
from app.routes.v1.user_role_route import _resolve_actor_context
from app.schemas.best_paper_schema import BestPaperSchema
from app.extensions import db
from app.utils.decorator import require_roles
from app.models.enumerations import Role, Status
from werkzeug.utils import secure_filename

from app.utils.services.mail import send_mail
from app.utils.services.sms import send_sms
from app.utils.model_utils import best_paper_utils

# Import utility functions
from app.utils.model_utils.best_paper_utils import (
    create_best_paper as create_best_paper_util,
    get_best_paper_by_id as get_best_paper_by_id_util,
    list_best_papers as list_best_papers_util,
    update_best_paper as update_best_paper_util,
    delete_best_paper as delete_best_paper_util,
    assign_verifier as assign_best_paper_verifier_util,
    remove_verifier as remove_best_paper_verifier_util,
)
from app.utils.model_utils.author_utils import (
    create_author as create_author_util,
    get_or_create_author as get_or_create_author_util,
)
from app.utils.model_utils.user_utils import (
    get_user_by_id as get_user_by_id_util,
    list_users as list_users_util,
)
from app.utils.model_utils.audit_log_utils import (
    create_audit_log as create_audit_log_util,
    record_event as record_event_util,
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# NOTE: Underlying model/schema still named BestPaper for now; outward API renamed to best_papers
best_paper_schema = BestPaperSchema()
best_papers_schema = BestPaperSchema(many=True)

def log_audit_event(event_type, user_id, details, ip_address=None, target_user_id=None):
    """Helper function to create audit logs with proper transaction handling"""
    try:
        # Create audit log without committing to avoid transaction issues
        create_audit_log_util(
            event=event_type,
            user_id=user_id,
            target_user_id=target_user_id,
            ip=ip_address,
            detail=json.dumps(details) if isinstance(details, dict) else details,
            actor_id=user_id,
            commit=False  # Don't commit here to avoid transaction issues
        )
        db.session.commit() # Commit only this specific operation
    except Exception as e:
        current_app.logger.error(f"Failed to create audit log: {str(e)}")
        try:
            db.session.rollback()
        except:
            pass  # Ignore rollback errors in error handling

def safe_commit():
    """Safely commit the database session with error handling"""
    try:
        db.session.commit()
        return True
    except Exception as e:
        current_app.logger.error(f"Database commit failed: {str(e)}")
        try:
            db.session.rollback()
        except:
            current_app.logger.error("Database rollback also failed")
        return False

@research_bp.route('/best-papers', methods=['POST'])
@jwt_required()
def create_best_paper():
    """Create a new Best Paper submission."""
    current_user_id = get_jwt_identity()
    user = None
    best_paper = None
    try:
        # Get the current user ID from JWT
        user = get_user_by_id_util(current_user_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="best_paper.create.failed",
                user_id=current_user_id,
                details={"error": error_msg},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        # Handle both JSON and multipart form-data (for file uploads)
        if request.is_json:
            data = request.get_json()
            full_paper_path = None
            forwarding_letter_path = None
        else:
            data_json = request.form.get('data')
            if not data_json:
                error_msg = "Request validation failed: No data provided in form request"
                log_audit_event(
                    event_type="best_paper.create.failed",
                    user_id=current_user_id,
                    details={"error": error_msg},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 40

            try:
                # Set created_by_id in data for schema load
                data = json.loads(data_json)
                data['created_by_id'] = current_user_id
            except json.JSONDecodeError:
                error_msg = "Request validation failed: Invalid JSON data provided"
                log_audit_event(
                    event_type="best_paper.create.failed",
                    user_id=current_user_id,
                    details={"error": error_msg},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 40

            # Handle file uploads
            full_paper_path = None
            forwarding_letter_path = None

            # Main best paper PDF (frontend key: bestpaper_pdf; maintain compatibility with abstract_pdf if provided)
            pdf_file = request.files.get('bestpaper_pdf') or request.files.get('abstract_pdf')
            if pdf_file:
                upload_folder = current_app.config.get('UPLOAD_FOLDER', 'uploads')
                os.makedirs(upload_folder, exist_ok=True)
                filename = secure_filename(pdf_file.filename)
                if not filename:
                    error_msg = "File validation failed: Invalid filename provided"
                    log_audit_event(
                        event_type="best_paper.create.failed",
                        user_id=current_user_id,
                        details={"error": error_msg, "filename": pdf_file.filename},
                        ip_address=request.remote_addr
                    )
                    return jsonify({"error": error_msg}), 400
                unique_filename = f"{uuid.uuid4().hex}_{filename}"
                file_path = os.path.join(upload_folder, unique_filename)
                pdf_file.save(file_path)
                current_app.logger.info(f"Best paper PDF saved to: {file_path}")
                # Store relative path like abstracts route (strip leading app/ if present)
                full_paper_path = file_path.replace("app/", "", 1)

            # Forwarding letter PDF (frontend key: forwarding_pdf)
            fwd_file = request.files.get('forwarding_pdf')
            if fwd_file:
                upload_folder = current_app.config.get('UPLOAD_FOLDER', 'uploads')
                os.makedirs(upload_folder, exist_ok=True)
                filename = secure_filename(fwd_file.filename)
                if not filename:
                    error_msg = "File validation failed: Invalid forwarding letter filename provided"
                    log_audit_event(
                        event_type="best_paper.create.failed",
                        user_id=current_user_id,
                        details={"error": error_msg, "filename": fwd_file.filename},
                        ip_address=request.remote_addr
                    )
                    return jsonify({"error": error_msg}), 40
                unique_filename = f"{uuid.uuid4().hex}_{filename}"
                file_path = os.path.join(upload_folder, unique_filename)
                fwd_file.save(file_path)
                current_app.logger.info(f"Forwarding letter PDF saved to: {file_path}")
                forwarding_letter_path = file_path.replace("app/", "", 1)

        # Check permissions
        if not (
                user.has_role(Role.ADMIN.value) or
                user.has_role(Role.SUPERADMIN.value) or
                user.has_role(Role.VERIFIER.value) or
                user.has_role(Role.COORDINATOR.value)
        ):
            # For non-admin users, ensure they can only create best papers for themselves
            pass  # This will be handled by setting created_by_id

        # Extract potential authors array (frontend sends single-author array)
        authors_data = data.pop('authors', []) or []

        # If client already provided author_id use it, otherwise create an Author from first authors entry
        author_id = data.get('author_id')
        if not author_id:
            if authors_data:
                a0 = authors_data[0] or {}
                author_name = a0.get('name', '').strip()
                if not author_name:
                    error_msg = "Validation failed: Author name is required to create a best paper"
                    log_audit_event(
                        event_type="best_paper.create.failed",
                        user_id=current_user_id,
                        details={"error": error_msg},
                        ip_address=request.remote_addr
                    )
                    return jsonify({"error": error_msg}), 400
                
                # Use get_or_create_author utility
                author, created = get_or_create_author_util(
                    name=author_name,
                    affiliation=a0.get('affiliation'),
                    email=a0.get('email'),
                    actor_id=current_user_id
                )
                author_id = str(author.id)
                current_app.logger.info(f"{'Created' if created else 'Found'} author {author.name} with ID {author_id} for best paper")
            else:
                error_msg = "Validation failed: Author information is required to create a best paper"
                log_audit_event(
                    event_type="best_paper.create.failed",
                    user_id=current_user_id,
                    details={"error": error_msg},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 40

        # Validate required fields
        if 'paper_category_id' not in data:
            error_msg = "Validation failed: paper_category_id is required to create a best paper"
            log_audit_event(
                event_type="best_paper.create.failed",
                user_id=current_user_id,
                details={"error": error_msg, "missing_field": "paper_category_id"},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400

        # Check if submission window is open for best papers
        cycle_id = data.get('cycle_id')
        if cycle_id:
            from app.utils.model_utils.cycle_utils import get_cycle_by_id as get_cycle_by_id_util
            cycle = get_cycle_by_id_util(cycle_id)
            if cycle:
                from app.models.Cycle import CyclePhase
                from app.utils.model_utils.cycle_utils import list_windows
                from datetime import date
                active_windows = list_windows(cycle_id=cycle_id, reference_date=date.today())
                best_paper_windows = [w for w in active_windows if w.phase == CyclePhase.BEST_PAPER_SUBMISSION]
                if not best_paper_windows:
                    error_msg = f"Submission validation failed: Best paper submissions are not allowed for cycle {cycle.name} at this time"
                    log_audit_event(
                        event_type="best_paper.create.failed",
                        user_id=current_user_id,
                        details={
                            "error": error_msg,
                            "cycle_id": cycle_id,
                            "cycle_name": cycle.name
                        },
                        ip_address=request.remote_addr
                    )
                    return jsonify({"error": error_msg}), 400

        # Inject resolved author_id and any uploaded file paths
        data['author_id'] = author_id
        if full_paper_path:
            data['full_paper_path'] = full_paper_path
        if forwarding_letter_path:
            data['forwarding_letter_path'] = forwarding_letter_path

        # Use utility function to create best paper
        best_paper = create_best_paper_util(
            actor_id=current_user_id,
            **data
        )
        
        # Log successful creation
        log_audit_event(
            event_type="best_paper.create.success",
            user_id=current_user_id,
            details={
                "best_paper_id": best_paper.id,
                "title": best_paper.title,
                "author_id": author_id,
                "has_pdf": bool(full_paper_path),
                "has_forwarding_letter": bool(forwarding_letter_path)
            },
            ip_address=request.remote_addr
        )
        
        # Send notification (SMS and Email) to the user who created the best paper
        send_sms(
            user.mobile, f"Your best paper(Oncology) id : {best_paper.id} has been created successfully and is pending submission for review."
        )

        send_mail(
            user.email,
            "Best Paper(Oncology) Created Successfully",
            f"Dear {user.username},\n\nYour best paper(Oncology) with ID {best_paper.id} has been created successfully and is pending submission for review.\n Details:\nTitle: {best_paper.title}\nBest Paper ID: {best_paper.id}\n\nBest regards,\nResearch Section,AIIMS"
        )

        return jsonify(best_paper_schema.dump(best_paper)), 201
    except ValueError as ve:
        # Handle specific validation errors from model constraints
        if "Submissions are allowed only during the CyclePhase.BEST_PAPER_SUBMISSION period" in str(ve):
            error_msg = "Submission validation failed: Best paper submissions are not allowed at this time. Please check the active cycle windows."
            log_audit_event(
                event_type="best_paper.create.failed",
                user_id=current_user_id if current_user_id else (user.id if user else None),
                details={"error": error_msg, "exception_type": "ValueError", "exception_message": str(ve)},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
        else:
            current_app.logger.exception("ValueError creating best paper")
            error_msg = f"Validation error occurred while creating best paper: {str(ve)}"
            log_audit_event(
                event_type="best_paper.create.failed",
                user_id=current_user_id if current_user_id else (user.id if user else None),
                details={"error": error_msg, "exception_type": "ValueError", "exception_message": str(ve)},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 40
    except Exception as e:
        current_app.logger.exception("Error creating best paper")
        error_msg = f"System error occurred while creating best paper: {str(e)}"
        
        # Try to log the failure, but handle transaction issues gracefully
        try:
            log_audit_event(
                event_type="best_paper.create.failed",
                user_id=current_user_id if current_user_id else (user.id if user else None),
                details={"error": error_msg, "exception_type": type(e).__name__, "exception_message": str(e)},
                ip_address=request.remote_addr
            )
        except:
            # If logging fails, at least log to app logger
            current_app.logger.error(f"Failed to log audit event for best paper creation failure: {str(e)}")
        
        # Perform a safe rollback
        try:
            db.session.rollback()
        except:
            current_app.logger.error("Database rollback failed during error handling")
        
        return jsonify({"error": error_msg}), 40


@research_bp.route('/best-papers/<best_paper_id>/pdf', methods=['GET'])
@jwt_required()
def get_best_paper_pdf(best_paper_id):
    current_user_id = get_jwt_identity()
    try:
        best_paper = get_best_paper_by_id_util(best_paper_id)
        if not best_paper:
            error_msg = f"Resource not found: Best paper with ID {best_paper_id} does not exist"
            log_audit_event(
                event_type="best_paper.pdf.access.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Best paper not found.")
        
        if not best_paper.full_paper_path:
            error_msg = f"File not found: No PDF uploaded for best paper ID {best_paper_id}"
            log_audit_event(
                event_type="best_paper.pdf.access.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id},
                ip_address=request.remote_addr
            )
            abort(404, description="No PDF uploaded for this best paper.")
        
        # Log successful access
        log_audit_event(
            event_type="best_paper.pdf.access.success",
            user_id=current_user_id,
            details={
                "best_paper_id": best_paper_id,
                "file_path": best_paper.full_paper_path
            },
            ip_address=request.remote_addr
        )
        
        return send_file(best_paper.full_paper_path, mimetype='application/pdf', as_attachment=False)
    except FileNotFoundError:
        error_msg = f"File access error: PDF file not found at path {best_paper.full_paper_path if 'best_paper' in locals() else 'unknown'}"
        log_audit_event(
            event_type="best_paper.pdf.access.failed",
            user_id=current_user_id,
            details={"error": error_msg, "best_paper_id": best_paper_id},
            ip_address=request.remote_addr
        )
        abort(404, description="PDF file not found.")
    except Exception as e:
        current_app.logger.exception("Error sending PDF file")
        error_msg = f"System error occurred while accessing PDF: {str(e)}"
        log_audit_event(
            event_type="best_paper.pdf.access.failed",
            user_id=current_user_id,
            details={"error": error_msg, "best_paper_id": best_paper_id},
            ip_address=request.remote_addr
        )
        abort(500, description="Internal server error.")


@research_bp.route('/best-papers/<best_paper_id>/forwarding_pdf', methods=['GET'])
@jwt_required()
def get_best_paper_forwarding_pdf(best_paper_id):
    current_user_id = get_jwt_identity()
    try:
        best_paper = get_best_paper_by_id_util(best_paper_id)
        if not best_paper:
            error_msg = f"Resource not found: Best paper with ID {best_paper_id} does not exist"
            log_audit_event(
                event_type="best_paper.forwarding_pdf.access.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Best paper not found.")
        
        if not best_paper.forwarding_letter_path:
            error_msg = f"File not found: No forwarding PDF uploaded for best paper ID {best_paper_id}"
            log_audit_event(
                event_type="best_paper.forwarding.pdf.access.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id},
                ip_address=request.remote_addr
            )
            abort(404, description="No forwarding PDF uploaded for this best paper.")
        
        # Log successful access
        log_audit_event(
            event_type="best_paper.forwarding.pdf.access.success",
            user_id=current_user_id,
            details={
                "best_paper_id": best_paper_id,
                "file_path": best_paper.forwarding_letter_path
            },
            ip_address=request.remote_addr
        )
        
        return send_file(best_paper.forwarding_letter_path, mimetype='application/pdf', as_attachment=False)
    except FileNotFoundError:
        error_msg = f"File access error: Forwarding PDF file not found at path {best_paper.forwarding_letter_path if 'best_paper' in locals() else 'unknown'}"
        log_audit_event(
            event_type="best_paper.forwarding.pdf.access.failed",
            user_id=current_user_id,
            details={"error": error_msg, "best_paper_id": best_paper_id},
            ip_address=request.remote_addr
        )
        abort(404, description="PDF file not found.")
    except Exception as e:
        current_app.logger.exception("Error sending PDF file")
        error_msg = f"System error occurred while accessing forwarding PDF: {str(e)}"
        log_audit_event(
            event_type="best_paper.forwarding.pdf.access.failed",
            user_id=current_user_id,
            details={"error": error_msg, "best_paper_id": best_paper_id},
            ip_address=request.remote_addr
        )
        abort(500, description="Internal server error.")

@research_bp.route('/best-papers', methods=['GET'])
@jwt_required()
def get_best_papers():
    """Get all Best Paper submissions with filtering and pagination support."""
    current_user_id = get_jwt_identity()
    try:
        # Get query parameters
        q = request.args.get('q', '').strip()
        verifiers = request.args.get('verifiers', '').strip()
        page = int(request.args.get('page', 1))
        status = request.args.get('status', '').strip().upper()
        page_size = int(request.args.get('page_size', 20))
        sort_by = request.args.get('sort_by', 'id')
        sort_dir = request.args.get('sort_dir', 'desc')
        verifier_filter = request.args.get('verifier', '').strip().lower() == 'true'
        
        # Validate page size
        page_size = min(page_size, 100)  # Limit max page size
        page = max(1, page)  # Ensure page is at least 1
        
        # Get current user for permissions
        user = get_user_by_id_util(current_user_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="best_paper.list.failed",
                user_id=current_user_id,
                details={"error": error_msg},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        # Build filters based on user permissions and query parameters
        filters = []
        
        # Apply search filter
        if q:
            q_int = q.isdigit() and int(q) or None
            search_filters = []
            
            # Search in title
            search_filters.append(BestPaper.title.ilike(f'%{q}%'))
            
            # Search in content if it exists
            if hasattr(BestPaper, 'content'):
                search_filters.append(BestPaper.content.ilike(f'%{q}%'))
            
            # Search by ID if q is numeric
            if q_int:
                search_filters.append(BestPaper.id == q_int)
            
            if search_filters:
                from sqlalchemy import or_
                filters.append(or_(*search_filters))
        
        # Apply status filter
        if status in ['PENDING', 'UNDER_REVIEW', 'ACCEPTED', 'REJECTED']:
            status_value = Status[status]
            filters.append(BestPaper.status == status_value)
        elif status:  # Invalid status provided
            error_msg = f"Validation failed: Invalid status '{status}'. Valid statuses are PENDING, UNDER_REVIEW, ACCEPTED, REJECTED"
            log_audit_event(
                event_type="best_paper.list.failed",
                user_id=current_user_id,
                details={"error": error_msg, "invalid_status": status},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 40
        
        # Apply permissions filter
        if not (
                user.has_role(Role.ADMIN.value) or
                user.has_role(Role.SUPERADMIN.value) or
                user.has_role(Role.VERIFIER.value) or
                user.has_role(Role.COORDINATOR.value)
        ):
            # Regular users can only see their own best papers
            filters.append(BestPaper.created_by_id == current_user_id)
        
        # Apply verifier filter (filter by current user if they are a verifier)
        if verifier_filter:
            # Only show best papers assigned to the current user
            from sqlalchemy import exists
            filters.append(
                exists().where(
                    (BestPaperVerifiers.best_paper_id == BestPaper.id) & 
                    (BestPaperVerifiers.user_id == current_user_id)
                )
            )
        
        # Apply verifiers filter
        if verifiers:
            if verifiers.lower() == 'yes':
                # Only best papers with assigned verifiers
                from sqlalchemy import exists
                filters.append(
                    exists().where(
                        (BestPaperVerifiers.best_paper_id == BestPaper.id)
                    )
                )
            elif verifiers.lower() == 'no':
                # Only best papers without assigned verifiers
                from sqlalchemy import not_, exists
                filters.append(
                    not_(exists().where(
                        (BestPaperVerifiers.best_paper_id == BestPaper.id)
                    ))
                )
            else:
                error_msg = f"Validation failed: Invalid verifiers parameter '{verifiers}'. Valid values are 'yes' or 'no'"
                log_audit_event(
                    event_type="best_paper.list.failed",
                    user_id=current_user_id,
                    details={"error": error_msg, "invalid_verifiers_param": verifiers},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 40
        
        # Apply sorting
        order_by = None
        if sort_by == 'title':
            order_by = BestPaper.title.asc() if sort_dir.lower() == 'asc' else BestPaper.title.desc()
        elif sort_by == 'created_at':
            order_by = BestPaper.created_at.asc() if sort_dir.lower() == 'asc' else BestPaper.created_at.desc()
        elif sort_by == 'id':
            order_by = BestPaper.id.asc() if sort_dir.lower() == 'asc' else BestPaper.id.desc()
        else: # invalid sort field
            error_msg = f"Validation failed: Invalid sort field '{sort_by}'. Valid fields are 'id', 'title', 'created_at'"
            log_audit_event(
                event_type="best_paper.list.failed",
                user_id=current_user_id,
                details={"error": error_msg, "invalid_sort_by": sort_by},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 40
        
        # Calculate offset
        offset = (page - 1) * page_size
        
        # Use utility function to list best papers with filters
        best_papers = list_best_papers_util(
            filters=filters,
            order_by=order_by,
            limit=page_size,
            offset=offset,
            eager=True,  # Load related data like author, verifiers
            actor_id=current_user_id
        )
        
        # Count total records (without pagination)
        from sqlalchemy import func
        total_query = db.session.query(func.count(BestPaper.id))
        for f in filters:
            total_query = total_query.filter(f)
        total = total_query.scalar()
        
        # Add verifiers count to each best paper
        best_papers_data = []
        for best_paper in best_papers:
            best_paper_dict = best_paper_schema.dump(best_paper)
            # Count verifiers assigned to this best paper
            verifiers_count = db.session.query(BestPaperVerifiers).filter_by(best_paper_id=best_paper.id).count()
            best_paper_dict['verifiers_count'] = verifiers_count
            best_papers_data.append(best_paper_dict)
        
        # Prepare response
        response = {
            'items': best_papers_data,
            'total': total,
            'page': page,
            'pages': (total + page_size - 1) // page_size,
            'page_size': page_size
        }
        
        # Log successful retrieval
        log_audit_event(
            event_type="best_paper.list.success",
            user_id=current_user_id,
            details={
                "filters_applied": bool(filters),
                "search_query": q if q else None,
                "status_filter": status if status else None,
                "verifier_filter": verifier_filter,
                "results_count": len(best_papers_data),
                "total_count": total,
                "page": page,
                "page_size": page_size
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(response), 200
    except Exception as e:
        current_app.logger.exception("Error listing best papers with parameters")
        error_msg = f"System error occurred while retrieving best papers: {str(e)}"
        log_audit_event(
            event_type="best_paper.list.failed",
            user_id=current_user_id,
            details={"error": error_msg, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400

@research_bp.route('/best-papers/<best_paper_id>', methods=['GET'])
@jwt_required()
def get_best_paper(best_paper_id):
    """Get a specific Best Paper submission."""
    current_user_id = get_jwt_identity()
    try:
        best_paper = get_best_paper_by_id_util(best_paper_id)
        if not best_paper:
            error_msg = f"Resource not found: Best paper with ID {best_paper_id} does not exist"
            log_audit_event(
                event_type="best_paper.get.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        data = best_paper_schema.dump(best_paper)
        # Add PDF URLs if available
        if best_paper.full_paper_path:
            data['pdf_url'] = f"/api/v1/research/best-papers/{best_paper_id}/pdf"
        if best_paper.forwarding_letter_path:
            data['forwarding_pdf_url'] = f"/api/v1/research/best-papers/{best_paper_id}/forwarding_pdf"
        
        # Log successful retrieval
        log_audit_event(
            event_type="best_paper.get.success",
            user_id=current_user_id,
            details={
                "best_paper_id": best_paper_id,
                "title": best_paper.title,
                "has_pdf": bool(best_paper.full_paper_path),
                "has_forwarding_letter": bool(best_paper.forwarding_letter_path)
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(data), 200
    except Exception as e:
        current_app.logger.exception("Error retrieving best paper")
        error_msg = f"System error occurred while retrieving best paper: {str(e)}"
        log_audit_event(
            event_type="best_paper.get.failed",
            user_id=current_user_id,
            details={"error": error_msg, "best_paper_id": best_paper_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400

@research_bp.route('/best-papers/<best_paper_id>', methods=['PUT'])
@jwt_required()
def update_best_paper(best_paper_id):
    """Update a Best Paper submission."""
    current_user_id = get_jwt_identity()
    best_paper = None
    try:
        best_paper = get_best_paper_by_id_util(best_paper_id)
        if not best_paper:
            error_msg = f"Resource not found: Best paper with ID {best_paper_id} does not exist"
            log_audit_event(
                event_type="best_paper.update.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        # Check if user is authorized to update this best paper
        user = get_user_by_id_util(current_user_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="best_paper.update.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        # In a real implementation, you might want to check if the user
        # is the author or has admin privileges
        if not (
                user.has_role(Role.ADMIN.value) or
                user.has_role(Role.SUPERADMIN.value) or
                user.has_role(Role.VERIFIER.value) or
                user.has_role(Role.COORDINATOR.value) or
                best_paper.created_by_id == current_user_id
        ):
            error_msg = f"Authorization failed: You are not authorized to update best paper ID {best_paper_id}"
            log_audit_event(
                event_type="best_paper.update.failed",
                user_id=current_user_id,
                details={
                    "error": error_msg, 
                    "best_paper_id": best_paper_id,
                    "user_role": user.role_associations[0].role.value if user.role_associations else "no_role",
                    "best_paper_creator_id": best_paper.created_by_id
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 403

        data = request.get_json()
        
        # Use utility function to update best paper
        updated_best_paper = update_best_paper_util(
            best_paper,
            actor_id=current_user_id,
            **data
        )
        
        # Log successful update
        log_audit_event(
            event_type="best_paper.update.success",
            user_id=current_user_id,
            details={
                "best_paper_id": best_paper_id,
                "updated_fields": list(data.keys()),
                "title": updated_best_paper.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(best_paper_schema.dump(updated_best_paper)), 200
    except Exception as e:
        current_app.logger.exception("Error updating best paper")
        error_msg = f"System error occurred while updating best paper: {str(e)}"
        log_audit_event(
            event_type="best_paper.update.failed",
            user_id=current_user_id,
            details={"error": error_msg, "best_paper_id": best_paper_id if best_paper else best_paper_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 40

@research_bp.route('/best-papers/<best_paper_id>', methods=['DELETE'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def delete_best_paper(best_paper_id):
    """Delete a Best Paper submission."""
    current_user_id = get_jwt_identity()
    best_paper = None
    try:
        best_paper = get_best_paper_by_id_util(best_paper_id)
        if not best_paper:
            error_msg = f"Resource not found: Best paper with ID {best_paper_id} does not exist"
            log_audit_event(
                event_type="best_paper.delete.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        # Use utility function to delete best paper
        delete_best_paper_util(
            best_paper_id,
            actor_id=current_user_id
        )
        
        # Log successful deletion
        log_audit_event(
            event_type="best_paper.delete.success",
            user_id=current_user_id,
            details={
                "best_paper_id": best_paper_id,
                "title": best_paper.title,
                "deleted_by": current_user_id
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Best Paper deleted"}), 200
    except Exception as e:
        current_app.logger.exception("Error deleting best paper")
        error_msg = f"System error occurred while deleting best paper: {str(e)}"
        log_audit_event(
            event_type="best_paper.delete.failed",
            user_id=current_user_id,
            details={"error": error_msg, "best_paper_id": best_paper_id if best_paper else best_paper_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 40

@research_bp.route('/best-papers/<best_paper_id>/submit', methods=['POST'])
@jwt_required()
def submit_best_paper(best_paper_id):
    """Submit a Best Paper for review."""
    current_user_id = get_jwt_identity()
    best_paper = None
    try:
        best_paper = get_best_paper_by_id_util(best_paper_id)
        if not best_paper:
            error_msg = f"Resource not found: Best paper with ID {best_paper_id} does not exist"
            log_audit_event(
                event_type="best_paper.submit.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        # Check if user is authorized to submit this best paper
        user = get_user_by_id_util(current_user_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="best_paper.submit.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        # Only the creator of the best paper or admin can submit it
        if not (
                best_paper.created_by_id == current_user_id or
                user.has_role(Role.ADMIN.value) or
                user.has_role(Role.SUPERADMIN.value)
        ):
            error_msg = f"Authorization failed: You are not authorized to submit best paper ID {best_paper_id} for review"
            log_audit_event(
                event_type="best_paper.submit.failed",
                user_id=current_user_id,
                details={
                    "error": error_msg, 
                    "best_paper_id": best_paper_id,
                    "user_role": user.role_associations[0].role.value if user.role_associations else "no_role",
                    "best_paper_creator_id": best_paper.created_by_id
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 403

        # Use utility function to update best paper status
        updated_best_paper = update_best_paper_util(
            best_paper,
            status=Status.UNDER_REVIEW,
            actor_id=current_user_id
        )
        
        # Log successful submission
        log_audit_event(
            event_type="best_paper.submit.success",
            user_id=current_user_id,
            details={
                "best_paper_id": best_paper_id,
                "new_status": "UNDER_REVIEW",
                "title": updated_best_paper.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Best Paper submitted for review"}), 200
    except Exception as e:
        current_app.logger.exception("Error submitting best paper")
        error_msg = f"System error occurred while submitting best paper: {str(e)}"
        log_audit_event(
            event_type="best_paper.submit.failed",
            user_id=current_user_id,
            details={"error": error_msg, "best_paper_id": best_paper_id if best_paper else best_paper_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400

@research_bp.route('/best-papers/status', methods=['GET'])
@jwt_required()
def get_best_paper_submission_status():
    """Get submission status of Best Paper submissions for the current user."""
    current_user_id = get_jwt_identity()
    try:
        user = get_user_by_id_util(current_user_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="best_paper.status.get.failed",
                user_id=current_user_id,
                details={"error": error_msg},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        # Build filters based on user permissions
        filters = []
        
        if user.has_role(Role.ADMIN.value) or user.has_role(Role.SUPERADMIN.value):
            # Admins can see all submissions
            pass
        elif user.has_role(Role.VERIFIER.value):
            # Verifiers can see submissions assigned to them
            from sqlalchemy import exists
            filters.append(
                exists().where(
                    (BestPaperVerifiers.best_paper_id == BestPaper.id) & 
                    (BestPaperVerifiers.user_id == current_user_id)
                )
            )
        else:
            # Regular users can only see their own submissions
            filters.append(BestPaper.created_by_id == current_user_id)

        # Count best papers by status
        pending_count = db.session.query(BestPaper).filter(
            *filters,
            BestPaper.status == Status.PENDING
        ).count()
        
        under_review_count = db.session.query(BestPaper).filter(
            *filters,
            BestPaper.status == Status.UNDER_REVIEW
        ).count()
        
        accepted_count = db.session.query(BestPaper).filter(
            *filters,
            BestPaper.status == Status.ACCEPTED
        ).count()
        
        rejected_count = db.session.query(BestPaper).filter(
            *filters,
            BestPaper.status == Status.REJECTED
        ).count()

        # Log successful retrieval
        log_audit_event(
            event_type="best_paper.status.get.success",
            user_id=current_user_id,
            details={
                "user_role": user.role_associations[0].role.value if user.role_associations else "no_role",
                "pending_count": pending_count,
                "under_review_count": under_review_count,
                "accepted_count": accepted_count,
                "rejected_count": rejected_count
            },
            ip_address=request.remote_addr
        )

        return jsonify({
            "pending": pending_count,
            "under_review": under_review_count,
            "accepted": accepted_count,
            "rejected": rejected_count
        }), 200
    except Exception as e:
        current_app.logger.exception("Error getting best paper submission status")
        error_msg = f"System error occurred while retrieving best paper status: {str(e)}"
        log_audit_event(
            event_type="best_paper.status.get.failed",
            user_id=current_user_id,
            details={"error": error_msg, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 40


# Verifier Management Routes for Best Papers

@research_bp.route('/best-papers/<best_paper_id>/verifiers/<user_id>', methods=['POST'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def assign_verifier_to_best_paper(best_paper_id, user_id):
    """Assign a verifier to a best paper."""
    current_user_id = get_jwt_identity()
    best_paper = None
    user = None
    try:
        # Check if best paper exists
        best_paper = get_best_paper_by_id_util(best_paper_id)
        if not best_paper:
            error_msg = f"Resource not found: Best paper with ID {best_paper_id} does not exist"
            log_audit_event(
                event_type="best_paper.verifier.assign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        
        # Check if user exists and is a verifier
        user = get_user_by_id_util(user_id)
        if not user:
            error_msg = f"Resource not found: User with ID {user_id} does not exist"
            log_audit_event(
                event_type="best_paper.verifier.assign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        if not user.has_role(Role.VERIFIER.value):
            error_msg = f"Validation failed: User with ID {user_id} is not a verifier"
            log_audit_event(
                event_type="best_paper.verifier.assign.failed",
                user_id=current_user_id,
                details={
                    "error": error_msg, 
                    "best_paper_id": best_paper_id, 
                    "verifier_id": user_id,
                    "user_role": user.role_associations[0].role.value if user.role_associations else "no_role"
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 40
        
        # Use utility function to assign verifier
        updated_best_paper = assign_best_paper_verifier_util(
            best_paper,
            user,
            actor_id=current_user_id
        )
        
        # Log successful assignment
        log_audit_event(
            event_type="best_paper.verifier.assign.success",
            user_id=current_user_id,
            details={
                "best_paper_id": best_paper_id,
                "verifier_id": user_id,
                "verifier_username": user.username,
                "title": updated_best_paper.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Verifier assigned successfully"}), 201
    except Exception as e:
        current_app.logger.exception("Error assigning verifier to best paper")
        error_msg = f"System error occurred while assigning verifier to best paper: {str(e)}"
        log_audit_event(
            event_type="best_paper.verifier.assign.failed",
            user_id=current_user_id,
            details={"error": error_msg, "best_paper_id": best_paper_id if best_paper else best_paper_id, "verifier_id": user_id if user else user_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/best-papers/<best_paper_id>/verifiers/<user_id>', methods=['DELETE'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def unassign_verifier_from_best_paper(best_paper_id, user_id):
    """Unassign a verifier from a best paper."""
    current_user_id = get_jwt_identity()
    best_paper = None
    user = None
    try:
        # Check if best paper exists
        best_paper = get_best_paper_by_id_util(best_paper_id)
        if not best_paper:
            error_msg = f"Resource not found: Best paper with ID {best_paper_id} does not exist"
            log_audit_event(
                event_type="best_paper.verifier.unassign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        
        # Check if user exists
        user = get_user_by_id_util(user_id)
        if not user:
            error_msg = f"Resource not found: User with ID {user_id} does not exist"
            log_audit_event(
                event_type="best_paper.verifier.unassign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        
        # Use utility function to remove verifier
        updated_best_paper = remove_best_paper_verifier_util(
            best_paper,
            user,
            actor_id=current_user_id
        )
        
        # Log successful unassignment
        log_audit_event(
            event_type="best_paper.verifier.unassign.success",
            user_id=current_user_id,
            details={
                "best_paper_id": best_paper_id,
                "verifier_id": user_id,
                "verifier_username": user.username,
                "title": updated_best_paper.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Verifier unassigned successfully"}), 200
    except Exception as e:
        current_app.logger.exception("Error unassigning verifier from best paper")
        error_msg = f"System error occurred while unassigning verifier from best paper: {str(e)}"
        log_audit_event(
            event_type="best_paper.verifier.unassign.failed",
            user_id=current_user_id,
            details={"error": error_msg, "best_paper_id": best_paper_id if best_paper else best_paper_id, "verifier_id": user_id if user else user_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/best-papers/<best_paper_id>/verifiers', methods=['GET'])
@jwt_required()
def get_verifiers_for_best_paper(best_paper_id):
    """Get all verifiers assigned to a best paper."""
    current_user_id = get_jwt_identity()
    best_paper = None
    try:
        # Check if best paper exists
        best_paper = get_best_paper_by_id_util(best_paper_id)
        if not best_paper:
            error_msg = f"Resource not found: Best paper with ID {best_paper_id} does not exist"
            log_audit_event(
                event_type="best_paper.verifiers.get.failed",
                user_id=current_user_id,
                details={"error": error_msg, "best_paper_id": best_paper_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        
        # Get all verifiers for this best paper using the relationship
        verifiers = best_paper.verifiers
        
        # Convert to simple dict format
        verifiers_data = []
        for verifier in verifiers:
            verifiers_data.append({
                'id': str(verifier.id),
                'username': verifier.username,
                'email': verifier.email,
                'employee_id': verifier.employee_id
            })
        
        # Log successful retrieval
        log_audit_event(
            event_type="best_paper.verifiers.get.success",
            user_id=current_user_id,
            details={
                "best_paper_id": best_paper_id,
                "verifiers_count": len(verifiers_data),
                "verifiers": [v['username'] for v in verifiers_data]
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(verifiers_data), 200
    except Exception as e:
        current_app.logger.exception("Error getting verifiers for best paper")
        error_msg = f"System error occurred while retrieving verifiers for best paper: {str(e)}"
        log_audit_event(
            event_type="best_paper.verifiers.get.failed",
            user_id=current_user_id,
            details={"error": error_msg, "best_paper_id": best_paper_id if best_paper else best_paper_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/verifiers/<user_id>/best-papers', methods=['GET'])
@jwt_required()
def get_best_papers_for_verifier(user_id):
    """Get all best papers assigned to a verifier."""
    current_user_id = get_jwt_identity()
    user = None
    try:
        # Check if user exists
        user = get_user_by_id_util(user_id)
        if not user:
            error_msg = f"Resource not found: User with ID {user_id} does not exist"
            log_audit_event(
                event_type="verifier.best_papers.get.failed",
                user_id=current_user_id,
                details={"error": error_msg, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        
        # Get all best papers assigned to this verifier using the relationship
        best_papers = user.best_papers_assigned  # This assumes there's a relationship defined
        
        # If the relationship doesn't exist, we need to query manually
        from sqlalchemy.orm import joinedload
        best_papers = db.session.query(BestPaper).join(
            BestPaperVerifiers, BestPaper.id == BestPaperVerifiers.best_paper_id
        ).filter(
            BestPaperVerifiers.user_id == user_id
        ).options(joinedload(BestPaper.author)).all()
        
        # Log successful retrieval
        log_audit_event(
            event_type="verifier.best_papers.get.success",
            user_id=current_user_id,
            details={
                "verifier_id": user_id,
                "best_papers_count": len(best_papers),
                "best_papers": [bp.title for bp in best_papers]
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(best_papers_schema.dump(best_papers)), 200
    except Exception as e:
        current_app.logger.exception("Error getting best papers for verifier")
        error_msg = f"System error occurred while retrieving best papers for verifier: {str(e)}"
        log_audit_event(
            event_type="verifier.best_papers.get.failed",
            user_id=current_user_id,
            details={"error": error_msg, "verifier_id": user_id if user else user_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/best-papers/bulk-assign-verifiers', methods=['POST'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def bulk_assign_verifiers_to_best_papers():
    """Bulk assign verifiers to multiple best papers."""
    current_user_id = get_jwt_identity()
    try:
        data = request.get_json()
        
        # Validate input
        if not data or 'best_paper_ids' not in data or 'user_ids' not in data:
            error_msg = "Request validation failed: Missing required fields 'best_paper_ids' or 'user_ids' in request body"
            log_audit_event(
                event_type="best_paper.verifiers.bulk_assign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "provided_fields": list(data.keys()) if data else []},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
        
        best_paper_ids = data['best_paper_ids']
        user_ids = data['user_ids']
        
        # Validate that all best papers and users exist
        best_papers = [get_best_paper_by_id_util(bpid) for bpid in best_paper_ids]
        users = [get_user_by_id_util(uid) for uid in user_ids]
        
        # Check if any best papers or users are None (not found)
        if None in best_papers:
            missing_best_papers = [bpid for i, bpid in enumerate(best_paper_ids) if best_papers[i] is None]
            error_msg = f"Resource validation failed: Best papers with IDs {missing_best_papers} do not exist"
            log_audit_event(
                event_type="best_paper.verifiers.bulk_assign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "missing_best_papers": missing_best_papers},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        if None in users:
            missing_users = [uid for i, uid in enumerate(user_ids) if users[i] is None]
            error_msg = f"Resource validation failed: Users with IDs {missing_users} do not exist"
            log_audit_event(
                event_type="best_paper.verifiers.bulk_assign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "missing_users": missing_users},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        
        # Check if all users are verifiers
        non_verifiers = [user for user in users if not user.has_role(Role.VERIFIER.value)]
        if non_verifiers:
            non_verifier_ids = [str(user.id) for user in non_verifiers]
            error_msg = f"Validation failed: Users with IDs {non_verifier_ids} are not verifiers"
            log_audit_event(
                event_type="best_paper.verifiers.bulk_assign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "non_verifier_ids": non_verifier_ids},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 40
        
        # Create assignments using utility functions
        assignments_created = 0
        for best_paper_id in best_paper_ids:
            for user_id in user_ids:
                # Get the best paper and user objects
                best_paper = get_best_paper_by_id_util(best_paper_id)
                user = get_user_by_id_util(user_id)
                
                # Check if already assigned
                existing_assignment = db.session.query(BestPaperVerifiers).filter_by(
                    best_paper_id=best_paper_id, user_id=user_id).first()
                
                if not existing_assignment:
                    # Use the utility function to assign verifier
                    assign_best_paper_verifier_util(best_paper, user, actor_id=current_user_id)
                    assignments_created += 1
        
        # Log successful bulk assignment
        log_audit_event(
            event_type="best_paper.verifiers.bulk_assign.success",
            user_id=current_user_id,
            details={
                "best_paper_ids": best_paper_ids,
                "user_ids": user_ids,
                "assignments_created": assignments_created,
                "total_possible_assignments": len(best_paper_ids) * len(user_ids)
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({
            "message": f"Successfully created {assignments_created} assignments",
            "assignments_created": assignments_created
        }), 201
    except Exception as e:
        current_app.logger.exception("Error in bulk assigning verifiers to best papers")
        error_msg = f"System error occurred during bulk assignment of verifiers: {str(e)}"
        log_audit_event(
            event_type="best_paper.verifiers.bulk_assign.failed",
            user_id=current_user_id,
            details={"error": error_msg, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 40


@research_bp.route('/best-papers/bulk-unassign-verifiers', methods=['POST'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def bulk_unassign_verifiers_from_best_papers():
    """Bulk unassign verifiers from multiple best papers."""
    current_user_id = get_jwt_identity()
    try:
        data = request.get_json()
        
        # Validate input
        if not data or 'best_paper_ids' not in data or 'user_ids' not in data:
            error_msg = "Request validation failed: Missing required fields 'best_paper_ids' or 'user_ids' in request body"
            log_audit_event(
                event_type="best_paper.verifiers.bulk_unassign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "provided_fields": list(data.keys()) if data else []},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
        
        best_paper_ids = data['best_paper_ids']
        user_ids = data['user_ids']
        
        # Delete assignments
        assignments_deleted = db.session.query(BestPaperVerifiers).filter(
            BestPaperVerifiers.best_paper_id.in_(best_paper_ids),
            BestPaperVerifiers.user_id.in_(user_ids)
        ).delete(synchronize_session=False)
        
        # Explicitly commit the transaction
        if not safe_commit():
            error_msg = "Database commit failed during bulk unassignment"
            log_audit_event(
                event_type="best_paper.verifiers.bulk_unassign.failed",
                user_id=current_user_id,
                details={"error": error_msg},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 500
        
        # Log successful bulk unassignment
        log_audit_event(
            event_type="best_paper.verifiers.bulk_unassign.success",
            user_id=current_user_id,
            details={
                "best_paper_ids": best_paper_ids,
                "user_ids": user_ids,
                "assignments_deleted": assignments_deleted
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({
            "message": f"Successfully deleted {assignments_deleted} assignments",
            "assignments_deleted": assignments_deleted
        }), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Error in bulk unassigning verifiers from best papers")
        error_msg = f"System error occurred during bulk unassignment of verifiers: {str(e)}"
        log_audit_event(
            event_type="best_paper.verifiers.bulk_unassign.failed",
            user_id=current_user_id,
            details={"error": error_msg, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


def sanitize_excel_value(value):
    """
    Convert to string and strip characters illegal in Excel.
    """
    if value is None:
        return ""

    # If it's not string, convert
    if not isinstance(value, str):
        value = str(value)

    # Remove illegal characters (ASCII control chars that Excel doesn't allow)
    value = ILLEGAL_CHARACTERS_RE.sub("", value)

    # Optional: also strip other weird control chars if you want
    # value = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", value)

    return value

def create_paper_excel(papers):
    """Create an Excel workbook with all papers data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Papers Master Sheet"

    # Define headers with more detailed information from the models
    headers = [
        "ID", "Paper Number", "Title", "Category", "Status",
        "Created By", "Created At", "Updated At", "PDF Path",
        "Review Phase", "Cycle Name", "Cycle Start Date", "Cycle End Date",
        "Authors", "Author Emails", "Affiliations"
    ]

    # Add headers to worksheet
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add paper data to worksheet
    for row_num, paper in enumerate(papers, 2):
        # Get author information
        authors_list = paper.author.name if paper.author else None
        author_emails_list = paper.author.email if paper.author else None
        affiliations_list = paper.author.affiliation if paper.author else None

        # Get creator name
        creator_name = paper.created_by.username if paper.created_by_id else "N/A"

        # Get cycle information
        cycle_name = paper.cycle.name if paper.cycle else "N/A"
        cycle_start_date = paper.cycle.start_date.isoformat() if paper.cycle and paper.cycle.start_date else "N/A"
        cycle_end_date = paper.cycle.end_date.isoformat() if paper.cycle and paper.cycle.end_date else "N/A"

        # Get category information
        category_name = paper.category.name if paper.category else "N/A"

        # Add row data
        row_data = [
            str(paper.id),  # Convert UUID to string
            paper.paper_number or "N/A",
            paper.title or "N/A",
            category_name or "N/A",
            paper.status.name if paper.status else "N/A",
            creator_name or "N/A",
            paper.created_at.isoformat() if paper.created_at else "N/A",
            paper.updated_at.isoformat() if paper.updated_at else "N/A",
            paper.complete_pdf or "N/A",
            paper.review_phase or "N/A",
            cycle_name or "N/A",
            cycle_start_date or "N/A",
            cycle_end_date or "N/A",
            authors_list or "N/A",
            author_emails_list or "N/A",
            affiliations_list or "N/A",
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num,
                           value=sanitize_excel_value(value))
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Limit max width to 50
        ws.column_dimensions[column_letter].width = adjusted_width

    return wb


@research_bp.route('/best-papers/export-excel', methods=['GET'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def export_papers_excel():
    """Export all papers to an Excel file."""
    actor_id, context = _resolve_actor_context("export_papers_excel")
    try:
        # Get all abstracts with related data
        abstracts = best_paper_utils.list_papers(
            actor_id=actor_id,
            context=context
        )

        # Create Excel workbook
        wb = create_paper_excel(abstracts)

        # Save to a temporary file
        from io import BytesIO
        from flask import send_file
        import tempfile
        import os

        # Create a temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()  # Close the file so it can be read by send_file

        # Log successful export
        log_audit_event(
            event_type="abstract.excel.export.success",
            user_id=actor_id,
            details={
                "exported_count": len(abstracts),
                "file_path": temp_file.name
            },
            ip_address=request.remote_addr
        )

        # Send the file to the user
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f"papers_master_{uuid.uuid4().hex[:8]}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as exc:
        current_app.logger.exception("Error exporting papers to Excel")
        error_msg = f"System error occurred while exporting papers to Excel: {str(exc)}"
        log_audit_event(
            event_type="paper.excel.export.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(
                exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/best-papers/export-pdf-zip', methods=['GET'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def export_papers_pdf_zip():
    """Export all award PDFs in a ZIP file organized by category, including an Excel summary."""
    actor_id, context = _resolve_actor_context("export_papers_pdf_zip")
    try:
        import zipfile
        import shutil
        from io import BytesIO
        import tempfile

        # Get all awards with related data
        papers = best_paper_utils.list_papers(
            actor_id=actor_id,
            context=context
        )

        # Base path where PDF files are stored
        base_path = current_app.config.get("UPLOAD_FOLDER", "uploads")

        # Create a temporary directory to organize files
        temp_dir = tempfile.mkdtemp()

        # Organize awards by category
        categories = {}
        for paper in papers:
            if paper.complete_pdf and paper.category:
                category_name = paper.category.name
                if category_name not in categories:
                    categories[category_name] = []
                categories[category_name].append(paper)
        # Create subdirectories for each category and copy PDF files
        for category_name, paper_list in categories.items():
            category_dir = os.path.join(temp_dir, category_name)
            os.makedirs(category_dir, exist_ok=True)

            for paper in paper_list:
                if paper.complete_pdf:
                    # Construct the full path from the base path
                    full_pdf_path = os.path.join(
                        base_path, os.path.basename(paper.complete_pdf))

                    # Verify the file exists before copying
                    if os.path.exists(full_pdf_path):
                        # Get the original filename
                        original_filename = os.path.basename(full_pdf_path)
                        # Create new filename with abstract ID to avoid conflicts
                        new_filename = f"{paper.id.hex}_{original_filename}"
                        destination_path = os.path.join(
                            category_dir, new_filename)

                        # Copy the PDF file to the category folder
                        shutil.copy2(full_pdf_path, destination_path)
                    else:
                        current_app.logger.warning(
                            f"PDF file not found at path: {full_pdf_path}")

        # Generate Excel file with all awards data using the existing function
        excel_wb = create_paper_excel(papers)

        # Save the Excel file to the temporary directory root
        excel_path = os.path.join(
            temp_dir, f"papers_summary_{uuid.uuid4().hex[:8]}.xlsx")
        excel_wb.save(excel_path)

        # Create a ZIP file from the organized directory structure
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    # Calculate the archive name (path inside the ZIP)
                    archive_name = os.path.relpath(file_path, temp_dir)
                    zip_file.write(file_path, archive_name)

        # Clean up the temporary directory
        shutil.rmtree(temp_dir)

        # Prepare the ZIP file for download
        zip_buffer.seek(0)

        # Log successful export
        log_audit_event(
            event_type="paper.pdf.excel.zip.export.success",
            user_id=actor_id,
            details={
                "exported_count": len(papers),
                "categories_count": len(categories),
                "categories": list(categories.keys())
            },
            ip_address=request.remote_addr
        )

        # Send the ZIP file to the user
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=f"papers_pdfs_with_excel_{uuid.uuid4().hex[:8]}.zip",
            mimetype='application/zip'
        )

    except Exception as exc:
        current_app.logger.exception("Error exporting papers PDFs to ZIP")
        error_msg = f"System error occurred while exporting papers PDFs to ZIP: {str(exc)}"
        log_audit_event(
            event_type="paper.pdf.excel.zip.export.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(
                exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/best-papers/export-with-pdfs', methods=['GET'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def export_papers_with_pdfs():
    """Export all papers with PDFs organized by category in a ZIP file, including an Excel summary."""
    actor_id, context = _resolve_actor_context("export_papers_with_pdfs")
    try:
        import zipfile
        import shutil
        from io import BytesIO
        import tempfile

        # Get all awards with related data
        papers = best_paper_utils.list_papers(
            actor_id=actor_id,
            context=context
        )

        # Base path where PDF files are stored
        base_path = current_app.config.get("UPLOAD_FOLDER", "uploads")

        # Create a temporary directory to organize files
        temp_dir = tempfile.mkdtemp()

        # Organize awards by category
        categories = {}
        for paper in papers:
            if paper.complete_pdf and paper.category:
                category_name = paper.category.name
                if category_name not in categories:
                    categories[category_name] = []
                categories[category_name].append(paper)

        # Create subdirectories for each category and copy PDF files
        for category_name, paper_list in categories.items():
            category_dir = os.path.join(temp_dir, category_name)
            os.makedirs(category_dir, exist_ok=True)

            for paper in paper_list:
                if paper.complete_pdf:
                    # Construct the full path from the base path
                    full_pdf_path = os.path.join(
                        base_path, os.path.basename(paper.complete_pdf))

                    # Verify the file exists before copying
                    if os.path.exists(full_pdf_path):
                        # Get the original filename
                        original_filename = os.path.basename(full_pdf_path)
                        # Create new filename with paper ID to avoid conflicts
                        new_filename = f"{paper.id.hex}_{original_filename}"
                        destination_path = os.path.join(
                            category_dir, new_filename)

                        # Copy the PDF file to the category folder
                        shutil.copy2(full_pdf_path, destination_path)
                    else:
                        current_app.logger.warning(
                            f"PDF file not found at path: {full_pdf_path}")

        # Generate Excel file with all awards data using the existing function
        excel_wb = create_paper_excel(papers)

        # Save the Excel file to the temporary directory root
        excel_path = os.path.join(
            temp_dir, f"papers_summary_{uuid.uuid4().hex[:8]}.xlsx")
        excel_wb.save(excel_path)

        # Create a ZIP file from the organized directory structure
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    # Calculate the archive name (path inside the ZIP)
                    archive_name = os.path.relpath(file_path, temp_dir)
                    zip_file.write(file_path, archive_name)

        # Clean up the temporary directory
        shutil.rmtree(temp_dir)

        # Prepare the ZIP file for download
        zip_buffer.seek(0)

        # Log successful export
        log_audit_event(
            event_type="award.pdf.excel.zip.export.success",
            user_id=actor_id,
            details={
                "exported_count": len(papers),
                "categories_count": len(categories),
                "categories": list(categories.keys())
            },
            ip_address=request.remote_addr
        )

        # Send the ZIP file to the user
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=f"papers_pdfs_with_excel_{uuid.uuid4().hex[:8]}.zip",
            mimetype='application/zip'
        )

    except Exception as exc:
        current_app.logger.exception("Error exporting papers with PDFs to ZIP")
        error_msg = f"System error occurred while exporting papers with PDFs to ZIP: {str(exc)}"
        log_audit_event(
            event_type="paper.pdf.excel.zip.export.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(
                exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400
