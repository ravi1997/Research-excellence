from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from flask import request, jsonify, current_app, abort, send_file
import json
import os
import uuid
from flask_jwt_extended import jwt_required, get_jwt_identity
from io import BytesIO
import openpyxl
from app.models.User import User
from app.routes.v1.audit_log_route import _resolve_actor_context
from app.routes.v1.research import research_bp
from app.models.Cycle import AwardVerifiers, Awards, Author, Category, PaperCategory, Cycle, GradingType, Grading, GradingFor
from app.schemas.awards_schema import AwardsSchema
from app.extensions import db
from app.utils.decorator import require_roles
from app.models.enumerations import Role, Status
from werkzeug.utils import secure_filename
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


from app.utils.model_utils import award_utils
from app.utils.model_utils import abstract_utils
from app.utils.services.mail import send_mail
from app.utils.services.sms import send_sms

# Import utility functions
from app.utils.model_utils.award_utils import (
    create_award as create_award_util,
    get_award_by_id as get_award_by_id_util,
    list_awards as list_awards_util,
    update_award as update_award_util,
    delete_award as delete_award_util,
    assign_verifier as assign_award_verifier_util,
    remove_verifier as remove_award_verifier_util,
)
from app.utils.model_utils.author_utils import (
    create_author as create_author_util,
    get_or_create_author as get_or_create_author_util,
)
from app.utils.model_utils.user_utils import (
    get_user_by_id as get_user_by_id_util,
    list_users as list_users_util,
)
from app.utils.model_utils.audit_log_utils import (
    create_audit_log as create_audit_log_util,
    record_event as record_event_util,
)

award_schema = AwardsSchema()
awards_schema = AwardsSchema(many=True)

def log_audit_event(event_type, user_id, details, ip_address=None, target_user_id=None):
    """Helper function to create audit logs with proper transaction handling"""
    try:
        # Create audit log without committing to avoid transaction issues
        create_audit_log_util(
            event=event_type,
            user_id=user_id,
            target_user_id=target_user_id,
            ip=ip_address,
            detail=json.dumps(details) if isinstance(details, dict) else details,
            actor_id=user_id,
            commit=False  # Don't commit here to avoid transaction issues
        )
        db.session.commit() # Commit only this specific operation
    except Exception as e:
        current_app.logger.error(f"Failed to create audit log: {str(e)}")
        try:
            db.session.rollback()
        except:
            pass  # Ignore rollback errors in error handling

def safe_commit():
    """Safely commit the database session with error handling"""
    try:
        db.session.commit()
        return True
    except Exception as e:
        current_app.logger.error(f"Database commit failed: {str(e)}")
        try:
            db.session.rollback()
        except:
            current_app.logger.error("Database rollback also failed")
        return False

@research_bp.route('/awards', methods=['POST'])
@jwt_required()
def create_award():
    """Create a new research award."""
    current_user_id = get_jwt_identity()
    user = None
    award = None
    try:
        # Get the current user ID from JWT
        user = get_user_by_id_util(current_user_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="award.create.failed",
                user_id=current_user_id,
                details={"error": error_msg},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        # Handle both JSON and multipart form-data (for file uploads)
        if request.is_json:
            data = request.get_json()
            full_paper_path = None
            forwarding_letter_path = None
        else:
            data_json = request.form.get('data')
            if not data_json:
                error_msg = "Request validation failed: No data provided in form request"
                log_audit_event(
                    event_type="award.create.failed",
                    user_id=current_user_id,
                    details={"error": error_msg},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 400

            try:
                # Set created_by_id in data for schema load
                data = json.loads(data_json)
                data['created_by_id'] = current_user_id
            except json.JSONDecodeError:
                error_msg = "Request validation failed: Invalid JSON data provided"
                log_audit_event(
                    event_type="award.create.failed",
                    user_id=current_user_id,
                    details={"error": error_msg},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 400

            # Handle file uploads
            full_paper_path = None
            forwarding_letter_path = None

            # Main award PDF (frontend key: award_pdf; maintain compatibility with abstract_pdf if provided)
            pdf_file = request.files.get('award_pdf') or request.files.get('abstract_pdf')
            if pdf_file:
                upload_folder = current_app.config.get('UPLOAD_FOLDER', 'uploads')
                os.makedirs(upload_folder, exist_ok=True)
                filename = secure_filename(pdf_file.filename)
                if not filename:
                    error_msg = "File validation failed: Invalid filename provided"
                    log_audit_event(
                        event_type="award.create.failed",
                        user_id=current_user_id,
                        details={"error": error_msg, "filename": pdf_file.filename},
                        ip_address=request.remote_addr
                    )
                    return jsonify({"error": error_msg}), 400
                unique_filename = f"{uuid.uuid4().hex}_{filename}"
                file_path = os.path.join(upload_folder, unique_filename)
                pdf_file.save(file_path)
                current_app.logger.info(f"Award PDF saved to: {file_path}")
                # Store relative path like abstracts route (strip leading app/ if present)
                full_paper_path = file_path.replace("app/", "", 1)

            # Forwarding letter PDF (frontend key: forwarding_pdf)
            fwd_file = request.files.get('forwarding_pdf')
            if fwd_file:
                upload_folder = current_app.config.get('UPLOAD_FOLDER', 'uploads')
                os.makedirs(upload_folder, exist_ok=True)
                filename = secure_filename(fwd_file.filename)
                if not filename:
                    error_msg = "File validation failed: Invalid forwarding letter filename provided"
                    log_audit_event(
                        event_type="award.create.failed",
                        user_id=current_user_id,
                        details={"error": error_msg, "filename": fwd_file.filename},
                        ip_address=request.remote_addr
                    )
                    return jsonify({"error": error_msg}), 400
                unique_filename = f"{uuid.uuid4().hex}_{filename}"
                file_path = os.path.join(upload_folder, unique_filename)
                fwd_file.save(file_path)
                current_app.logger.info(f"Forwarding letter PDF saved to: {file_path}")
                forwarding_letter_path = file_path.replace("app/", "", 1)

        # Check permissions
        if not (
                user.has_role(Role.ADMIN.value) or
                user.has_role(Role.SUPERADMIN.value) or
                user.has_role(Role.VERIFIER.value) or
                user.has_role(Role.COORDINATOR.value)
        ):
            # For non-admin users, ensure they can only create awards for themselves
            pass  # This will be handled by setting created_by_id

        # Extract potential authors array (frontend sends single-author array)
        authors_data = data.pop('authors', []) or []

        # If client already provided author_id use it, otherwise create an Author from first authors entry
        author_id = data.get('author_id')
        if not author_id:
            if authors_data:
                a0 = authors_data[0] or {}
                author_name = a0.get('name', '').strip()
                if not author_name:
                    error_msg = "Validation failed: Author name is required to create an award"
                    log_audit_event(
                        event_type="award.create.failed",
                        user_id=current_user_id,
                        details={"error": error_msg},
                        ip_address=request.remote_addr
                    )
                    return jsonify({"error": error_msg}), 400
                
                # Use get_or_create_author utility
                author, created = get_or_create_author_util(
                    name=author_name,
                    affiliation=a0.get('affiliation'),
                    email=a0.get('email'),
                    actor_id=current_user_id
                )
                author_id = str(author.id)
                current_app.logger.info(f"{'Created' if created else 'Found'} author {author.name} with ID {author_id} for award")
            else:
                error_msg = "Validation failed: Author information is required to create an award"
                log_audit_event(
                    event_type="award.create.failed",
                    user_id=current_user_id,
                    details={"error": error_msg},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 40

        # Validate required fields
        if 'paper_category_id' not in data:
            error_msg = "Validation failed: paper_category_id is required to create an award"
            log_audit_event(
                event_type="award.create.failed",
                user_id=current_user_id,
                details={"error": error_msg, "missing_field": "paper_category_id"},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400

        # Check if submission window is open for awards
        cycle_id = data.get('cycle_id')
        if cycle_id:
            from app.utils.model_utils.cycle_utils import get_cycle_by_id as get_cycle_by_id_util
            cycle = get_cycle_by_id_util(cycle_id)
            if cycle:
                from app.models.Cycle import CyclePhase
                from app.utils.model_utils.cycle_utils import list_windows
                from datetime import date
                active_windows = list_windows(cycle_id=cycle_id, reference_date=date.today())
                award_windows = [w for w in active_windows if w.phase == CyclePhase.AWARD_SUBMISSION]
                if not award_windows:
                    error_msg = f"Submission validation failed: Award submissions are not allowed for cycle {cycle.name} at this time"
                    log_audit_event(
                        event_type="award.create.failed",
                        user_id=current_user_id,
                        details={
                            "error": error_msg,
                            "cycle_id": cycle_id,
                            "cycle_name": cycle.name
                        },
                        ip_address=request.remote_addr
                    )
                    return jsonify({"error": error_msg}), 400

        # Inject resolved author_id and any uploaded file paths
        data['author_id'] = author_id
        if full_paper_path:
            data['full_paper_path'] = full_paper_path
        if forwarding_letter_path:
            data['forwarding_letter_path'] = forwarding_letter_path

        # Use utility function to create award
        award = create_award_util(
            actor_id=current_user_id,
            **data
        )
        
        # Log successful creation
        log_audit_event(
            event_type="award.create.success",
            user_id=current_user_id,
            details={
                "award_id": award.id,
                "title": award.title,
                "author_id": author_id,
                "has_pdf": bool(full_paper_path),
                "has_forwarding_letter": bool(forwarding_letter_path)
            },
            ip_address=request.remote_addr
        )
        
        # Send notification (SMS and Email) to the user who created the award
        send_sms(
            user.mobile, f"Your award id : {award.id} has been created successfully and is pending submission for review."
        )
        send_mail(
            user.email,
            "Award Created Successfully",
            f"Dear {user.username},\n\nYour award with ID {award.id} has been created successfully and is pending submission for review.\n Details:\nTitle: {award.title}\nAward ID: {award.id}\n\nBest regards,\nResearch Section,AIIMS"
        )        
        return jsonify(award_schema.dump(award)), 201
    except ValueError as ve:
        # Handle specific validation errors from model constraints
        if "Submissions are allowed only during the CyclePhase.AWARD_SUBMISSION period" in str(ve):
            error_msg = "Submission validation failed: Award submissions are not allowed at this time. Please check the active cycle windows."
            log_audit_event(
                event_type="award.create.failed",
                user_id=current_user_id if current_user_id else (user.id if user else None),
                details={"error": error_msg, "exception_type": "ValueError", "exception_message": str(ve)},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
        else:
            current_app.logger.exception("ValueError creating award")
            error_msg = f"Validation error occurred while creating award: {str(ve)}"
            log_audit_event(
                event_type="award.create.failed",
                user_id=current_user_id if current_user_id else (user.id if user else None),
                details={"error": error_msg, "exception_type": "ValueError", "exception_message": str(ve)},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
    except Exception as e:
        current_app.logger.exception("Error creating award")
        error_msg = f"System error occurred while creating award: {str(e)}"
        
        # Try to log the failure, but handle transaction issues gracefully
        try:
            log_audit_event(
                event_type="award.create.failed",
                user_id=current_user_id if current_user_id else (user.id if user else None),
                details={"error": error_msg, "exception_type": type(e).__name__, "exception_message": str(e)},
                ip_address=request.remote_addr
            )
        except:
            # If logging fails, at least log to app logger
            current_app.logger.error(f"Failed to log audit event for award creation failure: {str(e)}")
        
        # Perform a safe rollback
        try:
            db.session.rollback()
        except:
            current_app.logger.error("Database rollback failed during error handling")
        
        return jsonify({"error": error_msg}), 400


@research_bp.route('/awards/<abstract_id>/pdf', methods=['GET'])
@jwt_required()
def get_awards_pdf(abstract_id):
    current_user_id = get_jwt_identity()
    try:
        abstract = get_award_by_id_util(abstract_id)
        if not abstract:
            error_msg = f"Resource not found: Award with ID {abstract_id} does not exist"
            log_audit_event(
                event_type="award.pdf.access.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": abstract_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Award not found.")
        
        if not abstract.full_paper_path:
            error_msg = f"File not found: No PDF uploaded for award ID {abstract_id}"
            log_audit_event(
                event_type="award.pdf.access.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": abstract_id},
                ip_address=request.remote_addr
            )
            abort(404, description="No PDF uploaded for this award.")
        
        # Log successful access
        log_audit_event(
            event_type="award.pdf.access.success",
            user_id=current_user_id,
            details={
                "award_id": abstract_id,
                "file_path": abstract.full_paper_path
            },
            ip_address=request.remote_addr
        )
        
        return send_file(abstract.full_paper_path, mimetype='application/pdf', as_attachment=False)
    except FileNotFoundError:
        error_msg = f"File access error: PDF file not found at path {abstract.full_paper_path if 'abstract' in locals() else 'unknown'}"
        log_audit_event(
            event_type="award.pdf.access.failed",
            user_id=current_user_id,
            details={"error": error_msg, "award_id": abstract_id},
            ip_address=request.remote_addr
        )
        abort(404, description="PDF file not found.")
    except Exception as e:
        current_app.logger.exception("Error sending PDF file")
        error_msg = f"System error occurred while accessing PDF: {str(e)}"
        log_audit_event(
            event_type="award.pdf.access.failed",
            user_id=current_user_id,
            details={"error": error_msg, "award_id": abstract_id},
            ip_address=request.remote_addr
        )
        abort(500, description="Internal server error.")


@research_bp.route('/awards/<abstract_id>/forwarding_pdf', methods=['GET'])
@jwt_required()
def get_awards_forwarding_pdf(abstract_id):
    current_user_id = get_jwt_identity()
    try:
        abstract = get_award_by_id_util(abstract_id)
        if not abstract:
            error_msg = f"Resource not found: Award with ID {abstract_id} does not exist"
            log_audit_event(
                event_type="award.forwarding_pdf.access.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": abstract_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Award not found.")
        
        if not abstract.forwarding_letter_path:
            error_msg = f"File not found: No forwarding letter PDF uploaded for award ID {abstract_id}"
            log_audit_event(
                event_type="award.forwarding_pdf.access.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": abstract_id},
                ip_address=request.remote_addr
            )
            abort(404, description="No forwarding letter PDF uploaded for this award.")
        
        # Log successful access
        log_audit_event(
            event_type="award.forwarding_pdf.access.success",
            user_id=current_user_id,
            details={
                "award_id": abstract_id,
                "file_path": abstract.forwarding_letter_path
            },
            ip_address=request.remote_addr
        )
        
        return send_file(abstract.forwarding_letter_path, mimetype='application/pdf', as_attachment=False)
    except FileNotFoundError:
        error_msg = f"File access error: Forwarding letter PDF file not found at path {abstract.forwarding_letter_path if 'abstract' in locals() else 'unknown'}"
        log_audit_event(
            event_type="award.forwarding_pdf.access.failed",
            user_id=current_user_id,
            details={"error": error_msg, "award_id": abstract_id},
            ip_address=request.remote_addr
        )
        abort(404, description="PDF file not found.")
    except Exception as e:
        current_app.logger.exception("Error sending forwarding PDF file")
        error_msg = f"System error occurred while accessing forwarding PDF: {str(e)}"
        log_audit_event(
            event_type="award.forwarding_pdf.access.failed",
            user_id=current_user_id,
            details={"error": error_msg, "award_id": abstract_id},
            ip_address=request.remote_addr
        )
        abort(500, description="Internal server error.")


@research_bp.route('/awards', methods=['GET'])
@jwt_required()
def get_awards():
    """Get all research awards with filtering and pagination support."""
    current_user_id = get_jwt_identity()
    try:
        # Get query parameters
        q = request.args.get('q', '').strip()
        verifiers = request.args.get('verifiers', '').strip()
        page = int(request.args.get('page', 1))
        status = request.args.get('status', '').strip().upper()
        page_size = int(request.args.get('page_size', 20))
        sort_by = request.args.get('sort_by', 'id')
        sort_dir = request.args.get('sort_dir', 'desc')
        verifier_filter = request.args.get('verifier', '').strip().lower() == 'true'
        
        # Validate page size
        page_size = min(page_size, 100)  # Limit max page size
        page = max(1, page)  # Ensure page is at least 1
        
        # Get current user for permissions
        user = get_user_by_id_util(current_user_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="award.list.failed",
                user_id=current_user_id,
                details={"error": error_msg},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        # Build filters based on user permissions and query parameters
        filters = []
        
        # Apply search filter
        if q:
            q_int = q.isdigit() and int(q) or None
            search_filters = []
            
            # Search in title
            search_filters.append(Awards.title.ilike(f'%{q}%'))
            
            # Search in description if it exists
            if hasattr(Awards, 'description'):
                search_filters.append(Awards.description.ilike(f'%{q}%'))
            
            # Search by ID if q is numeric
            if q_int:
                search_filters.append(Awards.award_number == q_int)
            
            if search_filters:
                from sqlalchemy import or_
                filters.append(or_(*search_filters))
        
        # Apply status filter
        if status in ['PENDING', 'UNDER_REVIEW', 'ACCEPTED', 'REJECTED']:
            status_value = Status[status]
            filters.append(Awards.status == status_value)
        elif status:  # Invalid status provided
            error_msg = f"Validation failed: Invalid status '{status}'. Valid statuses are PENDING, UNDER_REVIEW, ACCEPTED, REJECTED"
            log_audit_event(
                event_type="award.list.failed",
                user_id=current_user_id,
                details={"error": error_msg, "invalid_status": status},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
        
        # Apply permissions filter
        if not (
                user.has_role(Role.ADMIN.value) or
                user.has_role(Role.SUPERADMIN.value) or
                user.has_role(Role.VERIFIER.value) or
                user.has_role(Role.COORDINATOR.value)
        ):
            # Regular users can only see their own awards
            filters.append(Awards.created_by_id == current_user_id)
        
        if user.has_role(Role.COORDINATOR.value):
            # Coordinators can see awards in their categories
            try:
                cat_ids = [
                    c.id for c in user.award_categories if c is not None]
            except Exception:
                cat_ids = []
            if cat_ids:
                filters.append(Awards.paper_category_id.in_(cat_ids))
            else:
                # If no categories assigned, coordinator sees no awards
                filters.append(Awards.id == None)

        # Apply verifier filter (filter by current user if they are a verifier)
        if verifier_filter:
            # Only show awards assigned to the current user
            from sqlalchemy import exists
            filters.append(
                exists().where(
                    (AwardVerifiers.award_id == Awards.id) & 
                    (AwardVerifiers.user_id == current_user_id)
                )
            )
        
        # Apply verifiers filter
        if verifiers:
            if verifiers.lower() == 'yes':
                # Only awards with assigned verifiers
                from sqlalchemy import exists
                filters.append(
                    exists().where(
                        (AwardVerifiers.award_id == Awards.id)
                    )
                )
            elif verifiers.lower() == 'no':
                # Only awards without assigned verifiers
                from sqlalchemy import not_, exists
                filters.append(
                    not_(exists().where(
                        (AwardVerifiers.award_id == Awards.id)
                    ))
                )
            else:
                error_msg = f"Validation failed: Invalid verifiers parameter '{verifiers}'. Valid values are 'yes' or 'no'"
                log_audit_event(
                    event_type="award.list.failed",
                    user_id=current_user_id,
                    details={"error": error_msg, "invalid_verifiers_param": verifiers},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 400
        
        # Apply sorting
        order_by = None
        if sort_by == 'title':
            order_by = Awards.title.asc() if sort_dir.lower() == 'asc' else Awards.title.desc()
        elif sort_by == 'created_at':
            order_by = Awards.created_at.asc() if sort_dir.lower() == 'asc' else Awards.created_at.desc()
        elif sort_by == 'id':
            order_by = Awards.id.asc() if sort_dir.lower() == 'asc' else Awards.id.desc()
        else:  # invalid sort field
            error_msg = f"Validation failed: Invalid sort field '{sort_by}'. Valid fields are 'id', 'title', 'created_at'"
            log_audit_event(
                event_type="award.list.failed",
                user_id=current_user_id,
                details={"error": error_msg, "invalid_sort_by": sort_by},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
        
        # Calculate offset
        offset = (page - 1) * page_size
        
        # Use utility function to list awards with filters
        awards = list_awards_util(
            filters=filters,
            order_by=order_by,
            limit=page_size,
            offset=offset,
            eager=True,  # Load related data like author, verifiers
            actor_id=current_user_id
        )
        
        # Count total records (without pagination)
        from sqlalchemy import func
        total_query = db.session.query(func.count(Awards.id))
        for f in filters:
            total_query = total_query.filter(f)
        total = total_query.scalar()
        
        # Add verifiers count to each award
        awards_data = []
        for award in awards:
            award_dict = award_schema.dump(award)
            # Count verifiers assigned to this award
            verifiers_count = db.session.query(AwardVerifiers).filter_by(award_id=award.id).count()
            award_dict['verifiers_count'] = verifiers_count
            awards_data.append(award_dict)
        
        # Prepare response
        response = {
            'items': awards_data,
            'total': total,
            'page': page,
            'pages': (total + page_size - 1) // page_size,
            'page_size': page_size
        }
        
        # Log successful retrieval
        log_audit_event(
            event_type="award.list.success",
            user_id=current_user_id,
            details={
                "filters_applied": bool(filters),
                "search_query": q if q else None,
                "status_filter": status if status else None,
                "verifier_filter": verifier_filter,
                "results_count": len(awards_data),
                "total_count": total,
                "page": page,
                "page_size": page_size
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(response), 200
    except Exception as e:
        current_app.logger.exception("Error listing awards with parameters")
        error_msg = f"System error occurred while retrieving awards: {str(e)}"
        log_audit_event(
            event_type="award.list.failed",
            user_id=current_user_id,
            details={"error": error_msg, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400

@research_bp.route('/awards/<award_id>', methods=['GET'])
@jwt_required()
def get_award(award_id):
    """Get a specific research award."""
    current_user_id = get_jwt_identity()
    try:
        award = get_award_by_id_util(award_id)
        if not award:
            error_msg = f"Resource not found: Award with ID {award_id} does not exist"
            log_audit_event(
                event_type="award.get.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": award_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        data = award_schema.dump(award)
        # Add PDF URLs if available
        if award.full_paper_path:
            data['pdf_url'] = f"/api/v1/research/awards/{award_id}/pdf"
        if award.forwarding_letter_path:
            data['forwarding_pdf_url'] = f"/api/v1/research/awards/{award_id}/forwarding_pdf"
        
        # Log successful retrieval
        log_audit_event(
            event_type="award.get.success",
            user_id=current_user_id,
            details={
                "award_id": award_id,
                "title": award.title,
                "has_pdf": bool(award.full_paper_path),
                "has_forwarding_letter": bool(award.forwarding_letter_path)
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(data), 200
    except Exception as e:
        current_app.logger.exception("Error retrieving award")
        error_msg = f"System error occurred while retrieving award: {str(e)}"
        log_audit_event(
            event_type="award.get.failed",
            user_id=current_user_id,
            details={"error": error_msg, "award_id": award_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400

@research_bp.route('/awards/<award_id>', methods=['PUT'])
@jwt_required()
def update_award(award_id):
    """Update a research award."""
    current_user_id = get_jwt_identity()
    award = None
    try:
        award = get_award_by_id_util(award_id)
        if not award:
            error_msg = f"Resource not found: Award with ID {award_id} does not exist"
            log_audit_event(
                event_type="award.update.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": award_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        # Check if user is authorized to update this award
        user = get_user_by_id_util(current_user_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="award.update.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": award_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        # In a real implementation, you might want to check if the user
        # is the author or has admin privileges
        if not (
                user.has_role(Role.ADMIN.value) or
                user.has_role(Role.SUPERADMIN.value) or
                user.has_role(Role.VERIFIER.value) or
                user.has_role(Role.COORDINATOR.value) or
                award.created_by_id == current_user_id
        ):
            error_msg = f"Authorization failed: You are not authorized to update award ID {award_id}"
            log_audit_event(
                event_type="award.update.failed",
                user_id=current_user_id,
                details={
                    "error": error_msg, 
                    "award_id": award_id,
                    "user_role": user.role_associations[0].role.value if user.role_associations else "no_role",
                    "award_creator_id": award.created_by_id
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 403

        data = request.get_json()
        if not data:
            error_msg = "Request validation failed: No JSON data provided in request body"
            log_audit_event(
                event_type="award.update.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": award_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
        
        # Use utility function to update award
        updated_award = update_award_util(
            award,
            actor_id=current_user_id,
            **data
        )
        
        # Log successful update
        log_audit_event(
            event_type="award.update.success",
            user_id=current_user_id,
            details={
                "award_id": award_id,
                "updated_fields": list(data.keys()),
                "title": updated_award.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(award_schema.dump(updated_award)), 200
    except Exception as e:
        current_app.logger.exception("Error updating award")
        error_msg = f"System error occurred while updating award: {str(e)}"
        log_audit_event(
            event_type="award.update.failed",
            user_id=current_user_id,
            details={"error": error_msg, "award_id": award_id if award else award_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400

@research_bp.route('/awards/<award_id>', methods=['DELETE'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def delete_award(award_id):
    """Delete a research award."""
    current_user_id = get_jwt_identity()
    award = None
    try:
        award = get_award_by_id_util(award_id)
        if not award:
            error_msg = f"Resource not found: Award with ID {award_id} does not exist"
            log_audit_event(
                event_type="award.delete.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": award_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        # Use utility function to delete award
        delete_award_util(
            award_id,
            actor_id=current_user_id
        )
        
        # Log successful deletion
        log_audit_event(
            event_type="award.delete.success",
            user_id=current_user_id,
            details={
                "award_id": award_id,
                "title": award.title,
                "deleted_by": current_user_id
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Award deleted"}), 200
    except Exception as e:
        current_app.logger.exception("Error deleting award")
        error_msg = f"System error occurred while deleting award: {str(e)}"
        log_audit_event(
            event_type="award.delete.failed",
            user_id=current_user_id,
            details={"error": error_msg, "award_id": award_id if award else award_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400

@research_bp.route('/awards/<award_id>/submit', methods=['POST'])
@jwt_required()
def submit_award(award_id):
    """Submit an award for review."""
    current_user_id = get_jwt_identity()
    award = None
    try:
        award = get_award_by_id_util(award_id)
        if not award:
            error_msg = f"Resource not found: Award with ID {award_id} does not exist"
            log_audit_event(
                event_type="award.submit.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": award_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        # Check if user is authorized to submit this award
        user = get_user_by_id_util(current_user_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="award.submit.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": award_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        # Only the creator of the award or admin can submit it
        if not (
                award.created_by_id == current_user_id or
                user.has_role(Role.ADMIN.value) or
                user.has_role(Role.SUPERADMIN.value)
        ):
            error_msg = f"Authorization failed: You are not authorized to submit award ID {award_id} for review"
            log_audit_event(
                event_type="award.submit.failed",
                user_id=current_user_id,
                details={
                    "error": error_msg, 
                    "award_id": award_id,
                    "user_role": user.role_associations[0].role.value if user.role_associations else "no_role",
                    "award_creator_id": award.created_by_id
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 403

        # Use utility function to update award status
        updated_award = update_award_util(
            award,
            status=Status.UNDER_REVIEW,
            actor_id=current_user_id
        )
        
        # Log successful submission
        log_audit_event(
            event_type="award.submit.success",
            user_id=current_user_id,
            details={
                "award_id": award_id,
                "new_status": "UNDER_REVIEW",
                "title": updated_award.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Award submitted for review"}), 200
    except Exception as e:
        current_app.logger.exception("Error submitting award")
        error_msg = f"System error occurred while submitting award: {str(e)}"
        log_audit_event(
            event_type="award.submit.failed",
            user_id=current_user_id,
            details={"error": error_msg, "award_id": award_id if award else award_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


# New endpoint to accept an abstract with grades
@research_bp.route('/awards/<award_id>/accept', methods=['POST'])
@jwt_required()
@require_roles(Role.VERIFIER.value, Role.COORDINATOR.value, Role.ADMIN.value, Role.SUPERADMIN.value)
def accept_award(award_id):
    """Accept an award after review."""
    actor_id, context = _resolve_actor_context("accept_award")
    award = None
    try:
        award = award_utils.get_award_by_id(
            award_id,
            actor_id=actor_id,
            context=context,
        )
        if award is None:
            error_msg = f"Resource not found: Award with ID {award_id} does not exist"
            log_audit_event(
                event_type="award.accept.failed",
                user_id=actor_id,
                details={"error": error_msg, "award_id": award_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Award not found.")

        # Check if user is authorized to accept this award (use utility lookup for consistency)
        user = get_user_by_id_util(actor_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="award.accept.failed",
                user_id=actor_id,
                details={"error": error_msg, "award_id": award_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        # Allow if user is admin, superadmin, or a coordinator for this award   
        privileged = any(
            user.has_role(role)
            for role in (Role.VERIFIER.value, Role.COORDINATOR.value, Role.ADMIN.value, Role.SUPERADMIN.value)
        )

        # Compare IDs as strings to avoid type mismatch between UUID and str
        if not privileged and not any(str(coordinator.id) == str(actor_id) for coordinator in award.coordinators):
            error_msg = f"Authorization failed: You are not authorized to accept award ID {award_id}"
            log_audit_event(
                event_type="award.accept.failed",
                user_id=actor_id,
                details={
                    "error": error_msg,
                    "award_id": award_id,
                    "user_id": actor_id,
                    "user_role": user.role_associations[0].role.value if user.role_associations else "no_role"
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 403

        # Check if all required grades have been submitted for the current phase
        if not award_utils.can_advance_to_next_phase(award, actor_id):
            error_msg = f"Cannot accept award: Not all required grades have been submitted for phase {award.review_phase}"
            log_audit_event(
                event_type="award.accept.failed",
                user_id=actor_id,
                details={
                    "error": error_msg,
                    "award_id": award_id,
                    "current_phase": award.review_phase
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400

        # Update the status to ACCEPTED
        award_utils.update_award(
            award,
            actor_id=actor_id,
            context=context,
            status=Status.ACCEPTED,
        )

        # Log successful acceptance
        log_audit_event(
            event_type="award.accept.success",
            user_id=actor_id,
            details={
                "award_id": award_id,
                "new_status": "ACCEPTED",
                "title": award.title
            },
            ip_address=request.remote_addr
        )

        return jsonify({"message": "Award accepted successfully"}), 200
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Error accepting award")
        error_msg = f"System error occurred while accepting award: {str(exc)}"
        log_audit_event(
            event_type="award.accept.failed",
            user_id=actor_id,
            details={"error": error_msg, "award_id": award_id if award else award_id,
                     "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400

@research_bp.route('/awards/status', methods=['GET'])
@jwt_required()
def get_award_submission_status():
    """Get submission status of awards for the current user."""
    current_user_id = get_jwt_identity()
    try:
        user = get_user_by_id_util(current_user_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="award.status.get.failed",
                user_id=current_user_id,
                details={"error": error_msg},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        # Build filters based on user permissions
        filters = []
        
        if user.has_role(Role.ADMIN.value) or user.has_role(Role.SUPERADMIN.value):
            # Admins can see all awards
            pass
        elif user.has_role(Role.VERIFIER.value):
            # Verifiers can see awards assigned to them
            from sqlalchemy import exists
            filters.append(
                exists().where(
                    (AwardVerifiers.award_id == Awards.id) & 
                    (AwardVerifiers.user_id == current_user_id)
                )
            )
        else:
            # Regular users can only see their own awards
            filters.append(Awards.created_by_id == current_user_id)

        # Count awards by status
        pending_awards = db.session.query(Awards).filter(
            *filters,
            Awards.status == Status.PENDING
        ).count()
        
        under_review_awards = db.session.query(Awards).filter(
            *filters,
            Awards.status == Status.UNDER_REVIEW
        ).count()
        
        accepted_awards = db.session.query(Awards).filter(
            *filters,
            Awards.status == Status.ACCEPTED
        ).count()
        
        rejected_awards = db.session.query(Awards).filter(
            *filters,
            Awards.status == Status.REJECTED
        ).count()

        # Log successful retrieval
        log_audit_event(
            event_type="award.status.get.success",
            user_id=current_user_id,
            details={
                "user_role": user.role_associations[0].role.value if user.role_associations else "no_role",
                "pending_count": pending_awards,
                "under_review_count": under_review_awards,
                "accepted_count": accepted_awards,
                "rejected_count": rejected_awards
            },
            ip_address=request.remote_addr
        )

        return jsonify({
            "pending": pending_awards,
            "under_review": under_review_awards,
            "accepted": accepted_awards,
            "rejected": rejected_awards
        }), 200
    except Exception as e:
        current_app.logger.exception("Error getting award submission status")
        error_msg = f"System error occurred while retrieving award status: {str(e)}"
        log_audit_event(
            event_type="award.status.get.failed",
            user_id=current_user_id,
            details={"error": error_msg, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


# Verifier Management Routes for Awards

@research_bp.route('/awards/<award_id>/verifiers/<user_id>', methods=['POST'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def assign_verifier_to_award(award_id, user_id):
    """Assign a verifier to an award."""
    current_user_id = get_jwt_identity()
    award = None
    user = None
    try:
        # Check if award exists
        award = get_award_by_id_util(award_id)
        if not award:
            error_msg = f"Resource not found: Award with ID {award_id} does not exist"
            log_audit_event(
                event_type="award.verifier.assign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": award_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        
        # Check if user exists and is a verifier
        user = get_user_by_id_util(user_id)
        if not user:
            error_msg = f"Resource not found: User with ID {user_id} does not exist"
            log_audit_event(
                event_type="award.verifier.assign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": award_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        if not user.has_role(Role.VERIFIER.value):
            error_msg = f"Validation failed: User with ID {user_id} is not a verifier"
            log_audit_event(
                event_type="award.verifier.assign.failed",
                user_id=current_user_id,
                details={
                    "error": error_msg, 
                    "award_id": award_id, 
                    "verifier_id": user_id,
                    "user_role": user.role_associations[0].role.value if user.role_associations else "no_role"
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
        
        # Use utility function to assign verifier
        updated_award = assign_award_verifier_util(
            award,
            user,
            actor_id=current_user_id
        )
        
        # Log successful assignment
        log_audit_event(
            event_type="award.verifier.assign.success",
            user_id=current_user_id,
            details={
                "award_id": award_id,
                "verifier_id": user_id,
                "verifier_username": user.username,
                "title": updated_award.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Verifier assigned successfully"}), 201
    except Exception as e:
        current_app.logger.exception("Error assigning verifier to award")
        error_msg = f"System error occurred while assigning verifier to award: {str(e)}"
        log_audit_event(
            event_type="award.verifier.assign.failed",
            user_id=current_user_id,
            details={"error": error_msg, "award_id": award_id if award else award_id, "verifier_id": user_id if user else user_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/awards/<award_id>/verifiers/<user_id>', methods=['DELETE'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def unassign_verifier_from_award(award_id, user_id):
    """Unassign a verifier from an award."""
    current_user_id = get_jwt_identity()
    award = None
    user = None
    try:
        # Check if award exists
        award = get_award_by_id_util(award_id)
        if not award:
            error_msg = f"Resource not found: Award with ID {award_id} does not exist"
            log_audit_event(
                event_type="award.verifier.unassign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": award_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        
        # Check if user exists
        user = get_user_by_id_util(user_id)
        if not user:
            error_msg = f"Resource not found: User with ID {user_id} does not exist"
            log_audit_event(
                event_type="award.verifier.unassign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": award_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        
        # Use utility function to remove verifier
        updated_award = remove_award_verifier_util(
            award,
            user,
            actor_id=current_user_id
        )
        
        # Log successful unassignment
        log_audit_event(
            event_type="award.verifier.unassign.success",
            user_id=current_user_id,
            details={
                "award_id": award_id,
                "verifier_id": user_id,
                "verifier_username": user.username,
                "title": updated_award.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Verifier unassigned successfully"}), 200
    except Exception as e:
        current_app.logger.exception("Error unassigning verifier from award")
        error_msg = f"System error occurred while unassigning verifier from award: {str(e)}"
        log_audit_event(
            event_type="award.verifier.unassign.failed",
            user_id=current_user_id,
            details={"error": error_msg, "award_id": award_id if award else award_id, "verifier_id": user_id if user else user_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/awards/<award_id>/verifiers', methods=['GET'])
@jwt_required()
def get_verifiers_for_award(award_id):
    """Get all verifiers assigned to an award."""
    current_user_id = get_jwt_identity()
    award = None
    try:
        # Check if award exists
        award = get_award_by_id_util(award_id)
        if not award:
            error_msg = f"Resource not found: Award with ID {award_id} does not exist"
            log_audit_event(
                event_type="award.verifiers.get.failed",
                user_id=current_user_id,
                details={"error": error_msg, "award_id": award_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        
        # Get all verifiers for this award using the relationship
        verifiers = award.verifiers
        
        # Convert to simple dict format
        verifiers_data = []
        for verifier in verifiers:
            verifiers_data.append({
                'id': str(verifier.id),
                'username': verifier.username,
                'email': verifier.email,
                'employee_id': verifier.employee_id
            })
        
        # Log successful retrieval
        log_audit_event(
            event_type="award.verifiers.get.success",
            user_id=current_user_id,
            details={
                "award_id": award_id,
                "verifiers_count": len(verifiers_data),
                "verifiers": [v['username'] for v in verifiers_data]
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(verifiers_data), 200
    except Exception as e:
        current_app.logger.exception("Error getting verifiers for award")
        error_msg = f"System error occurred while retrieving verifiers for award: {str(e)}"
        log_audit_event(
            event_type="award.verifiers.get.failed",
            user_id=current_user_id,
            details={"error": error_msg, "award_id": award_id if award else award_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/verifiers/<user_id>/awards', methods=['GET'])
@jwt_required()
def get_awards_for_verifier(user_id):
    """Get all awards assigned to a verifier."""
    current_user_id = get_jwt_identity()
    user = None
    try:
        # Check if user exists
        user = get_user_by_id_util(user_id)
        if not user:
            error_msg = f"Resource not found: User with ID {user_id} does not exist"
            log_audit_event(
                event_type="verifier.awards.get.failed",
                user_id=current_user_id,
                details={"error": error_msg, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        
        # Get all awards assigned to this verifier using the relationship
        # (model relationship is named `awards_to_verify` on `User`).
        awards = getattr(user, 'awards_to_verify', None)

        # If the relationship is not present or empty (e.g., not loaded), fall back to a direct query
        if not awards:
            from sqlalchemy.orm import joinedload
            awards = db.session.query(Awards).join(
                AwardVerifiers, Awards.id == AwardVerifiers.award_id
            ).filter(
                AwardVerifiers.user_id == user_id
            ).options(joinedload(Awards.author)).all()
        
        # Log successful retrieval
        log_audit_event(
            event_type="verifier.awards.get.success",
            user_id=current_user_id,
            details={
                "verifier_id": user_id,
                "awards_count": len(awards),
                "awards": [award.title for award in awards]
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(awards_schema.dump(awards)), 200
    except Exception as e:
        current_app.logger.exception("Error getting awards for verifier")
        error_msg = f"System error occurred while retrieving awards for verifier: {str(e)}"
        log_audit_event(
            event_type="verifier.awards.get.failed",
            user_id=current_user_id,
            details={"error": error_msg, "verifier_id": user_id if user else user_id, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/awards/bulk-assign-verifiers', methods=['POST'])
@jwt_required()
@require_roles(Role.COORDINATOR.value, Role.ADMIN.value, Role.SUPERADMIN.value)
def bulk_assign_verifiers_to_awards():
    """Bulk assign verifiers to multiple awards."""
    current_user_id = get_jwt_identity()
    try:
        data = request.get_json()
        
        # Validate input
        if not data or 'award_ids' not in data or 'user_ids' not in data:
            error_msg = "Request validation failed: Missing required fields 'award_ids' or 'user_ids' in request body"
            log_audit_event(
                event_type="award.verifiers.bulk_assign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "provided_fields": list(data.keys()) if data else []},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
        
        award_ids = data['award_ids']
        user_ids = data['user_ids']
        
        # Validate that all awards and users exist
        awards = [get_award_by_id_util(aid) for aid in award_ids]
        users = [get_user_by_id_util(uid) for uid in user_ids]
        
        # Check if any awards or users are None (not found)
        if None in awards:
            missing_awards = [aid for i, aid in enumerate(award_ids) if awards[i] is None]
            error_msg = f"Resource validation failed: Awards with IDs {missing_awards} do not exist"
            log_audit_event(
                event_type="award.verifiers.bulk_assign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "missing_awards": missing_awards},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        if None in users:
            missing_users = [uid for i, uid in enumerate(user_ids) if users[i] is None]
            error_msg = f"Resource validation failed: Users with IDs {missing_users} do not exist"
            log_audit_event(
                event_type="award.verifiers.bulk_assign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "missing_users": missing_users},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
        
        # Check if all users are verifiers
        non_verifiers = [user for user in users if not user.has_role(Role.VERIFIER.value)]
        if non_verifiers:
            non_verifier_ids = [str(user.id) for user in non_verifiers]
            error_msg = f"Validation failed: Users with IDs {non_verifier_ids} are not verifiers"
            log_audit_event(
                event_type="award.verifiers.bulk_assign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "non_verifier_ids": non_verifier_ids},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
        
        # Create assignments using utility functions
        assignments_created = 0
        for award_id in award_ids:
            for user_id in user_ids:
                # Get the award and user objects
                award = get_award_by_id_util(award_id)
                user = get_user_by_id_util(user_id)
                
                # Check if already assigned
                existing_assignment = db.session.query(AwardVerifiers).filter_by(
                    award_id=award_id, user_id=user_id).first()
                
                if not existing_assignment:
                    # Use the utility function to assign verifier
                    assign_award_verifier_util(award, user, actor_id=current_user_id)
                    assignments_created += 1
        
        # Log successful bulk assignment
        log_audit_event(
            event_type="award.verifiers.bulk_assign.success",
            user_id=current_user_id,
            details={
                "award_ids": award_ids,
                "user_ids": user_ids,
                "assignments_created": assignments_created,
                "total_possible_assignments": len(award_ids) * len(user_ids)
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({
            "message": f"Successfully created {assignments_created} assignments",
            "assignments_created": assignments_created
        }), 201
    except Exception as e:
        current_app.logger.exception("Error in bulk assigning verifiers to awards")
        error_msg = f"System error occurred during bulk assignment of verifiers: {str(e)}"
        log_audit_event(
            event_type="award.verifiers.bulk_assign.failed",
            user_id=current_user_id,
            details={"error": error_msg, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/awards/bulk-unassign-verifiers', methods=['POST'])
@jwt_required()
@require_roles(Role.COORDINATOR.value, Role.ADMIN.value, Role.SUPERADMIN.value)
def bulk_unassign_verifiers_from_awards():
    """Bulk unassign verifiers from multiple awards."""
    current_user_id = get_jwt_identity()
    try:
        data = request.get_json()
        
        # Validate input
        if not data or 'award_ids' not in data or 'user_ids' not in data:
            error_msg = "Request validation failed: Missing required fields 'award_ids' or 'user_ids' in request body"
            log_audit_event(
                event_type="award.verifiers.bulk_unassign.failed",
                user_id=current_user_id,
                details={"error": error_msg, "provided_fields": list(data.keys()) if data else []},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
        
        award_ids = data['award_ids']
        user_ids = data['user_ids']
        
        # Delete assignments
        assignments_deleted = db.session.query(AwardVerifiers).filter(
            AwardVerifiers.award_id.in_(award_ids),
            AwardVerifiers.user_id.in_(user_ids)
        ).delete(synchronize_session=False)
        
        # Explicitly commit the transaction
        if not safe_commit():
            error_msg = "Database commit failed during bulk unassignment"
            log_audit_event(
                event_type="award.verifiers.bulk_unassign.failed",
                user_id=current_user_id,
                details={"error": error_msg},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 500
        
        # Log successful bulk unassignment
        log_audit_event(
            event_type="award.verifiers.bulk_unassign.success",
            user_id=current_user_id,
            details={
                "award_ids": award_ids,
                "user_ids": user_ids,
                "assignments_deleted": assignments_deleted
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({
            "message": f"Successfully deleted {assignments_deleted} assignments",
            "assignments_deleted": assignments_deleted
        }), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("Error in bulk unassigning verifiers from awards")
        error_msg = f"System error occurred during bulk unassignment of verifiers: {str(e)}"
        log_audit_event(
            event_type="award.verifiers.bulk_unassign.failed",
            user_id=current_user_id,
            details={"error": error_msg, "exception_type": type(e).__name__, "exception_message": str(e)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


import re

def sanitize_excel_value(value):
    """
    Convert to string and strip characters illegal in Excel.
    """
    if value is None:
        return ""

    # If it's not string, convert
    if not isinstance(value, str):
        value = str(value)

    # Remove illegal characters (ASCII control chars that Excel doesn't allow)
    value = ILLEGAL_CHARACTERS_RE.sub("", value)

    # Optional: also strip other weird control chars if you want
    # value = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", value)

    return value

def create_awards_excel(awards):
    """Create an Excel workbook with all awards data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Awards Master Sheet"
    
    # Define headers with more detailed information from the models
    headers = [
        "ID", "Award Number", "Title", "Category", "Status",
        "Created By", "Created At", "Updated At", "PDF Path",
        "Review Phase", "Cycle Name", "Cycle Start Date", "Cycle End Date", 
        "Authors", "Author Emails", "Affiliations"
    ]
    
    # Add headers to worksheet
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add award data to worksheet
    for row_num, award in enumerate(awards, 2):
        # Get author information
        authors_list = award.author.name if award.author else None
        author_emails_list = award.author.email if award.author else None
        affiliations_list = award.author.affiliation if award.author else None

        # Get creator name
        creator_name = award.created_by.username if award.created_by_id else "N/A"

        # Get cycle information
        cycle_name = award.cycle.name if award.cycle else "N/A"
        cycle_start_date = award.cycle.start_date.isoformat() if award.cycle and award.cycle.start_date else "N/A"
        cycle_end_date = award.cycle.end_date.isoformat() if award.cycle and award.cycle.end_date else "N/A"

        # Get category information
        category_name = award.category.name if award.category else "N/A"

        # Add row data
        row_data = [
            str(award.id),  # Convert UUID to string
            award.award_number or "N/A",
            award.title or "N/A",
            category_name or "N/A",
            award.status.name if award.status else "N/A",
            creator_name or "N/A",
            award.created_at.isoformat() if award.created_at else "N/A",
            award.updated_at.isoformat() if award.updated_at else "N/A",
            award.complete_pdf or "N/A",
            award.review_phase or "N/A",
            cycle_name or "N/A",
            cycle_start_date or "N/A",
            cycle_end_date or "N/A",
            authors_list or "N/A",
            author_emails_list or "N/A",
            affiliations_list or "N/A",
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num,
                           value=sanitize_excel_value(value))
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Limit max width to 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb


@research_bp.route('/awards/export-with-grades', methods=['GET'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value, Role.COORDINATOR.value)
def export_awards_with_grades():
    """Export awards with their grades to Excel.

    Query parameters (optional): grading_type_id, cycle_id
    Produces one row per (award, grader) with grading-type columns.
    """
    actor_id, context = _resolve_actor_context("export_awards_with_grades")
    try:
        grading_type_id = request.args.get('grading_type_id', '').strip()
        cycle_id = request.args.get('cycle_id', '').strip()

        # Validate grading_type_id if provided
        if grading_type_id:
            try:
                uuid.UUID(grading_type_id)
            except ValueError:
                error_msg = f"Validation failed: Invalid grading_type_id '{grading_type_id}'"
                log_audit_event(
                    event_type="award.grades.export.failed",
                    user_id=actor_id,
                    details={"error": error_msg},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 400
            grading_type = GradingType.query.get(grading_type_id)
            if not grading_type:
                error_msg = f"Resource not found: Grading type with ID {grading_type_id} does not exist"
                log_audit_event(
                    event_type="award.grades.export.failed",
                    user_id=actor_id,
                    details={"error": error_msg, "grading_type_id": grading_type_id},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 404

        # Validate cycle_id if provided
        if cycle_id:
            try:
                uuid.UUID(cycle_id)
            except ValueError:
                cycle_id = ''

        # Get awards
        awards = award_utils.list_awards(actor_id=actor_id, context=context)

        # Coordinator filter: only awards in user's award_categories
        user = get_user_by_id_util(actor_id)
        if user and user.has_role(Role.COORDINATOR.value):
            try:
                cat_ids = [c.id for c in user.award_categories if c is not None]
            except Exception:
                cat_ids = []
            if cat_ids:
                awards = [a for a in awards if a.paper_category_id in cat_ids]
            else:
                awards = []

        # Cycle filter
        if cycle_id:
            awards = [a for a in awards if str(a.cycle_id) == cycle_id]

        # Build grading type columns for AWARD
        grading_types = GradingType.query.filter(GradingType.grading_for == GradingFor.AWARD).order_by(GradingType.criteria).all()

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Award Grades"

        base_headers = [
            "Award ID", "Award Number", "Title", "Category", "Cycle",
            "Status", "Created At", "Graded By"
        ]
        gt_headers = [gt.criteria for gt in grading_types]
        headers = base_headers + gt_headers

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Bulk fetch grades
        award_ids = [a.id for a in awards]
        gt_ids = [gt.id for gt in grading_types]

        grades = []
        if award_ids and gt_ids:
            grades = Grading.query.filter(
                Grading.award_id.in_(award_ids),
                Grading.grading_type_id.in_(gt_ids)
            ).all()

        # Map (award_id, grader_id, grading_type_id) -> score
        grade_map = {}
        graders_by_award = {}
        for g in grades:
            aid = str(g.award_id)
            gid = str(g.graded_by_id) if g.graded_by_id else "Unknown"
            tid = str(g.grading_type_id)
            grade_map[(aid, gid, tid)] = g.score
            graders_by_award.setdefault(aid, set()).add(gid)

        # Grader names
        grader_map = {}
        if grades:
            grader_ids = {str(g.graded_by_id) for g in grades if g.graded_by_id}
            if grader_ids:
                graders = User.query.filter(User.id.in_(grader_ids)).all()
                grader_map = {str(u.id): u.username for u in graders}

        # Fill rows
        row_num = 2
        for award in awards:
            aid = str(award.id)
            grader_ids = graders_by_award.get(aid, set())
            if not grader_ids:
                grader_ids = [None]
            for gid in sorted(grader_ids):
                grader_name = grader_map.get(gid, "Unknown") if gid else "No grades"
                row_values = [
                    aid,
                    award.award_number or "N/A",
                    award.title or "N/A",
                    award.category.name if award.category else "N/A",
                    award.cycle.name if award.cycle else "N/A",
                    award.status.name if award.status else "N/A",
                    award.created_at.isoformat() if award.created_at else "N/A",
                    grader_name
                ]
                for gt in grading_types:
                    key = (aid, gid if gid else "Unknown", str(gt.id))
                    score = grade_map.get(key)
                    row_values.append(score if score is not None else "N/A")

                for col_num, value in enumerate(row_values, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                row_num += 1

        # Auto-adjust widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.freeze_panes = "A2"

        # Convert to bytes and send
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        log_audit_event(
            event_type="award.grades.export.success",
            user_id=actor_id,
            details={
                "grading_type_id": grading_type_id if grading_type_id else "all",
                "cycle_filter": cycle_id if cycle_id else "all",
                "awards_count": len(awards)
            },
            ip_address=request.remote_addr
        )

        return send_file(
            output,
            download_name=f"awards_grades_{uuid.uuid4().hex[:8]}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as exc:
        current_app.logger.exception("Error exporting awards with grades")
        error_msg = f"System error occurred while exporting awards with grades: {str(exc)}"
        log_audit_event(
            event_type="award.grades.export.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 500




@research_bp.route('/awards/export-excel', methods=['GET'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def export_awards_to_excel():
    """Export all awards to an Excel file."""
    actor_id, context = _resolve_actor_context("export_awards_excel")
    try:
        # Get all awards with related data
        awards = award_utils.list_awards(
            actor_id=actor_id,
            context=context
        )
        
        # Create Excel workbook
        wb = create_awards_excel(awards)
        
        # Save to a temporary file
        from io import BytesIO
        from flask import send_file
        import tempfile
        import os
        
        # Create a temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()  # Close the file so it can be read by send_file

        # Log successful export
        log_audit_event(
            event_type="award.excel.export.success",
            user_id=actor_id,
            details={
                "exported_count": len(awards),
                "file_path": temp_file.name
            },
            ip_address=request.remote_addr
        )

        # Send the file to the user
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f"awards_master_{uuid.uuid4().hex[:8]}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as exc:
        current_app.logger.exception("Error exporting awards to Excel")
        error_msg = f"System error occurred while exporting awards to Excel: {str(exc)}"
        log_audit_event(
            event_type="award.excel.export.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/awards/export-pdf-zip', methods=['GET'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def export_awards_pdf_zip():
    """Export all award PDFs in a ZIP file organized by category, including an Excel summary."""
    actor_id, context = _resolve_actor_context("export_awards_pdf_zip")
    try:
        import zipfile
        import shutil
        from io import BytesIO
        import tempfile

        # Get all awards with related data
        awards = award_utils.list_awards(
            actor_id=actor_id,
            context=context
        )
        
        # Base path where PDF files are stored
        base_path = current_app.config.get("UPLOAD_FOLDER", "uploads")
        
        # Create a temporary directory to organize files
        temp_dir = tempfile.mkdtemp()
        
        # Organize awards by category
        categories = {}
        for award in awards:
            if award.complete_pdf and award.category:
                category_name = award.category.name
                if category_name not in categories:
                    categories[category_name] = []
                categories[category_name].append(award)

        # Create subdirectories for each category and copy PDF files
        for category_name, award_list in categories.items():
            category_dir = os.path.join(temp_dir, category_name)
            os.makedirs(category_dir, exist_ok=True)

            for award in award_list:
                if award.complete_pdf:
                    # Construct the full path from the base path
                    full_pdf_path = os.path.join(base_path, os.path.basename(award.complete_pdf))

                    # Verify the file exists before copying
                    if os.path.exists(full_pdf_path):
                        # Get the original filename
                        original_filename = os.path.basename(full_pdf_path)
                        # Create new filename with abstract ID to avoid conflicts
                        new_filename = f"{award.id.hex}_{original_filename}"
                        destination_path = os.path.join(category_dir, new_filename)
                        
                        # Copy the PDF file to the category folder
                        shutil.copy2(full_pdf_path, destination_path)
                    else:
                        current_app.logger.warning(f"PDF file not found at path: {full_pdf_path}")

        # Generate Excel file with all awards data using the existing function
        excel_wb = create_awards_excel(awards)

        # Save the Excel file to the temporary directory root
        excel_path = os.path.join(temp_dir, f"awards_summary_{uuid.uuid4().hex[:8]}.xlsx")
        excel_wb.save(excel_path)
        
        # Create a ZIP file from the organized directory structure
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    # Calculate the archive name (path inside the ZIP)
                    archive_name = os.path.relpath(file_path, temp_dir)
                    zip_file.write(file_path, archive_name)
        
        # Clean up the temporary directory
        shutil.rmtree(temp_dir)
        
        # Prepare the ZIP file for download
        zip_buffer.seek(0)
        
        # Log successful export
        log_audit_event(
            event_type="award.pdf.excel.zip.export.success",
            user_id=actor_id,
            details={
                "exported_count": len(awards),
                "categories_count": len(categories),
                "categories": list(categories.keys())
            },
            ip_address=request.remote_addr
        )
        
        # Send the ZIP file to the user
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=f"awards_pdfs_with_excel_{uuid.uuid4().hex[:8]}.zip",
            mimetype='application/zip'
        )
        
    except Exception as exc:
        current_app.logger.exception("Error exporting awards PDFs to ZIP")
        error_msg = f"System error occurred while exporting awards PDFs to ZIP: {str(exc)}"
        log_audit_event(
            event_type="award.pdf.excel.zip.export.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400

@research_bp.route('/awards/export-with-pdfs', methods=['GET'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def export_awards_with_pdfs():
    """Export all awards with PDFs organized by category in a ZIP file, including an Excel summary."""
    actor_id, context = _resolve_actor_context("export_awards_with_pdfs")
    try:
        import zipfile
        import shutil
        from io import BytesIO
        import tempfile

        # Get all awards with related data
        awards = award_utils.list_awards(
            actor_id=actor_id,
            context=context
        )
        
        # Base path where PDF files are stored
        base_path = current_app.config.get("UPLOAD_FOLDER", "uploads")
        
        # Create a temporary directory to organize files
        temp_dir = tempfile.mkdtemp()
        
        # Organize awards by category
        categories = {}
        for award in awards:
            if award.complete_pdf and award.category:
                category_name = award.category.name
                if category_name not in categories:
                    categories[category_name] = []
                categories[category_name].append(award)
        
        # Create subdirectories for each category and copy PDF files
        for category_name, award_list in categories.items():
            category_dir = os.path.join(temp_dir, category_name)
            os.makedirs(category_dir, exist_ok=True)

            for award in award_list:
                if award.complete_pdf:
                    # Construct the full path from the base path
                    full_pdf_path = os.path.join(base_path, os.path.basename(award.complete_pdf))

                    # Verify the file exists before copying
                    if os.path.exists(full_pdf_path):
                        # Get the original filename
                        original_filename = os.path.basename(full_pdf_path)
                        # Create new filename with award ID to avoid conflicts
                        new_filename = f"{award.id.hex}_{original_filename}"
                        destination_path = os.path.join(category_dir, new_filename)
                        
                        # Copy the PDF file to the category folder
                        shutil.copy2(full_pdf_path, destination_path)
                    else:
                        current_app.logger.warning(f"PDF file not found at path: {full_pdf_path}")

        # Generate Excel file with all awards data using the existing function
        excel_wb = create_awards_excel(awards)

        # Save the Excel file to the temporary directory root
        excel_path = os.path.join(temp_dir, f"awards_summary_{uuid.uuid4().hex[:8]}.xlsx")
        excel_wb.save(excel_path)
        
        # Create a ZIP file from the organized directory structure
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    # Calculate the archive name (path inside the ZIP)
                    archive_name = os.path.relpath(file_path, temp_dir)
                    zip_file.write(file_path, archive_name)
        
        # Clean up the temporary directory
        shutil.rmtree(temp_dir)
        
        # Prepare the ZIP file for download
        zip_buffer.seek(0)
        
        # Log successful export
        log_audit_event(
            event_type="award.pdf.excel.zip.export.success",
            user_id=actor_id,
            details={
                "exported_count": len(awards),
                "categories_count": len(categories),
                "categories": list(categories.keys())
            },
            ip_address=request.remote_addr
        )
        
        # Send the ZIP file to the user
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=f"awards_pdfs_with_excel_{uuid.uuid4().hex[:8]}.zip",
            mimetype='application/zip'
        )
        
    except Exception as exc:
        current_app.logger.exception("Error exporting awards with PDFs to ZIP")
        error_msg = f"System error occurred while exporting awards with PDFs to ZIP: {str(exc)}"
        log_audit_event(
            event_type="award.pdf.excel.zip.export.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400
