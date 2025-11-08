import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file, abort
# Route to serve the PDF file for an abstract

import json
import uuid
from typing import Dict, Optional, Tuple

from flask import request, jsonify, current_app
from flask_jwt_extended import get_jwt, get_jwt_identity, jwt_required
from sqlalchemy import or_

from app.extensions import db
from app.models.Cycle import (
    AbstractAuthors,
    Abstracts,
    Author,
    Category,
)
from app.models.Token import Token
from app.models.User import User
from app.models.enumerations import Role, Status
from app.routes.v1.research import research_bp
from app.schemas.abstract_schema import AbstractSchema
from app.utils.decorator import require_roles
from app.utils.model_utils import abstract_utils, audit_log_utils, token_utils
from app.utils.model_utils import author_utils
from app.utils.services.mail import send_mail
from app.utils.services.sms import send_sms
from app.models.Cycle import CyclePhase
from app.utils.model_utils.cycle_utils import get_cycle_by_id as get_cycle_by_id_util
from app.utils.model_utils.cycle_utils import list_windows

abstract_schema = AbstractSchema()
abstracts_schema = AbstractSchema(many=True)


def log_audit_event(event_type, user_id, details, ip_address=None, target_user_id=None):
    """Helper function to create audit logs with proper transaction handling"""
    try:
        # Create audit log without committing to avoid transaction issues
        audit_log_utils.create_audit_log(
            event=event_type,
            user_id=user_id,
            target_user_id=target_user_id,
            ip=ip_address,
            detail=json.dumps(details) if isinstance(details, dict) else details,
            actor_id=user_id,
            commit=False  # Don't commit here to avoid transaction issues
        )
        db.session.commit() # Commit only this specific operation
    except Exception as e:
        current_app.logger.error(f"Failed to create audit log: {str(e)}")
        try:
            db.session.rollback()
        except:
            pass  # Ignore rollback errors in error handling


def safe_commit():
    """Safely commit the database session with error handling"""
    try:
        db.session.commit()
        return True
    except Exception as e:
        current_app.logger.error(f"Database commit failed: {str(e)}")
        try:
            db.session.rollback()
        except:
            current_app.logger.error("Database rollback also failed")
        return False


def _resolve_actor_context(action: str) -> Tuple[Optional[str], Dict[str, object]]:
    """
    Resolve the acting user and build a context payload that reuses the shared
    token utilities so every route benefits from consistent logging.
    """

    actor_identity = get_jwt_identity()
    actor_id = str(actor_identity) if actor_identity is not None else None
    jwt_payload = get_jwt()
    token_jti: Optional[str] = jwt_payload.get("jti") if jwt_payload else None

    filters = [Token.jti == token_jti] if token_jti else [Token.id == 0]
    tokens = token_utils.list_tokens(
        filters=filters,
        actor_id=actor_id,
        context={"route": action, "token_jti": token_jti or "none"},
    )

    context: Dict[str, object] = {"route": action}
    if actor_id:
        context["actor_id"] = actor_id
    if token_jti:
        context["token_jti"] = token_jti
    if tokens:
        context["token_record_id"] = tokens[0].id

    return actor_id, context

@research_bp.route('/abstracts', methods=['POST'])
@jwt_required()
@require_roles(Role.USER.value,Role.ADMIN.value, Role.SUPERADMIN.value)
def create_abstract():
    """Create a new research abstract."""
    actor_id, context = _resolve_actor_context("create_abstract")
    if actor_id is None:
        error_msg = "Authentication failed: Unable to resolve actor identity"
        log_audit_event(
            event_type="abstract.create.failed",
            user_id=actor_id,
            details={"error": error_msg},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 404

    user = None
    abstract = None
    pdf_path: Optional[str] = None

    try:
        user = User.query.get(actor_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="abstract.create.failed",
                user_id=actor_id,
                details={"error": error_msg},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        if request.is_json:
            payload = request.get_json() or {}
            if not isinstance(payload, dict):
                error_msg = "Request validation failed: Invalid payload provided"
                log_audit_event(
                    event_type="abstract.create.failed",
                    user_id=actor_id,
                    details={"error": error_msg},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 400
        else:
            data_json = request.form.get("data")
            if not data_json:
                error_msg = "Request validation failed: No data provided in form request"
                log_audit_event(
                    event_type="abstract.create.failed",
                    user_id=actor_id,
                    details={"error": error_msg},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 400
            try:
                payload = json.loads(data_json)
            except json.JSONDecodeError:
                error_msg = "Request validation failed: Invalid JSON data provided"
                log_audit_event(
                    event_type="abstract.create.failed",
                    user_id=actor_id,
                    details={"error": error_msg},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 400

            pdf_file = request.files.get("abstract_pdf")
            if pdf_file:
                upload_folder = current_app.config.get("UPLOAD_FOLDER", "uploads")
                import os
                from werkzeug.utils import secure_filename

                os.makedirs(upload_folder, exist_ok=True)
                filename = secure_filename(pdf_file.filename)
                if not filename:
                    error_msg = "File validation failed: Invalid filename provided"
                    log_audit_event(
                        event_type="abstract.create.failed",
                        user_id=actor_id,
                        details={"error": error_msg, "filename": pdf_file.filename},
                        ip_address=request.remote_addr
                    )
                    return jsonify({"error": error_msg}), 400
                unique_filename = f"{uuid.uuid4().hex}_{filename}"
                file_path = os.path.join(upload_folder, unique_filename)
                pdf_file.save(file_path)
                current_app.logger.info("PDF file saved to: %s", file_path)
                pdf_path = file_path.replace("app/", "", 1)

        authors_data = payload.pop("authors", []) or []
        payload["created_by_id"] = actor_id
        if pdf_path:
            payload["pdf_path"] = pdf_path

        # Check if submission window is open for abstracts
        cycle_id = payload.get('cycle_id')
        if cycle_id:
            cycle = get_cycle_by_id_util(cycle_id)
            if cycle:
                from datetime import date
                active_windows = list_windows(cycle_id=cycle_id, reference_date=date.today())
                abstract_windows = [
                    w for w in active_windows if w.phase == CyclePhase.ABSTRACT_SUBMISSION]
                if not abstract_windows:
                    error_msg = f"Submission validation failed: Abstract submissions are not allowed for cycle {cycle.name} at this time"
                    log_audit_event(
                        event_type="abstract.create.failed",
                        user_id=actor_id,
                        details={
                            "error": error_msg,
                            "cycle_id": cycle_id,
                            "cycle_name": cycle.name
                        },
                        ip_address=request.remote_addr
                    )
                    return jsonify({"error": error_msg}), 400

        abstract = abstract_utils.create_abstract(
            commit=False,
            actor_id=actor_id,
            context=context,
            **payload,
        )

        db.session.flush()

        if authors_data:
            for index, author_data in enumerate(authors_data):
                author = author_utils.create_author(
                    commit=False,
                    actor_id=actor_id,
                    context={**context, "author_index": index},
                    name=author_data.get("name", ""),
                    email=author_data.get("email"),
                    affiliation=author_data.get("affiliation"),
                    is_presenter=author_data.get("is_presenter", False),
                    is_corresponding=author_data.get("is_corresponding", False),
                )
                db.session.flush()
                db.session.add(
                    AbstractAuthors(
                        abstract_id=abstract.id,
                        author_id=author.id,
                        author_order=index,
                    )
                )
                current_app.logger.info(
                    "Added author %s with ID %s to abstract %s",
                    author.name,
                    author.id,
                    abstract.id,
                )

        db.session.commit()

        if user:
            if getattr(user, "mobile", None):
                send_sms(
                    user.mobile,
                    f"Your abstract id : {abstract.id} has been created successfully and is pending submission for review.",
                )
            if getattr(user, "email", None):
                send_mail(
                    user.email,
                    "Abstract Created Successfully",
                    f"Dear {user.username},\n\nYour abstract with ID {abstract.id} has been created successfully and is pending submission for review.\n Details:\nTitle: {abstract.title}\nAbstract ID: {abstract.id}\n\nBest regards,\nResearch Section,AIIMS",
                )

        # Log successful creation
        log_audit_event(
            event_type="abstract.create.success",
            user_id=actor_id,
            details={
                "abstract_id": abstract.id,
                "title": abstract.title,
                "has_pdf": bool(pdf_path),
                "author_count": len(authors_data)
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(abstract_schema.dump(abstract)), 201
    except ValueError as ve:
        # Handle specific validation errors from model constraints
        if "Submissions are allowed only during the CyclePhase.SUBMISSION period" in str(ve):
            error_msg = "Submission validation failed: Abstract submissions are not allowed at this time. Please check the active cycle windows."
            log_audit_event(
                event_type="abstract.create.failed",
                user_id=actor_id if actor_id else (user.id if user else None),
                details={"error": error_msg, "exception_type": "ValueError", "exception_message": str(ve)},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
        else:
            current_app.logger.exception("ValueError creating abstract")
            error_msg = f"Validation error occurred while creating abstract: {str(ve)}"
            log_audit_event(
                event_type="abstract.create.failed",
                user_id=actor_id if actor_id else (user.id if user else None),
                details={"error": error_msg, "exception_type": "ValueError", "exception_message": str(ve)},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Error creating abstract")
        error_msg = f"System error occurred while creating abstract: {str(exc)}"
        
        # Try to log the failure, but handle transaction issues gracefully
        try:
            log_audit_event(
                event_type="abstract.create.failed",
                user_id=actor_id if actor_id else (user.id if user else None),
                details={"error": error_msg, "exception_type": type(exc).__name__, "exception_message": str(exc)},
                ip_address=request.remote_addr
            )
        except:
            # If logging fails, at least log to app logger
            current_app.logger.error(f"Failed to log audit event for abstract creation failure: {str(exc)}")
        
        return jsonify({"error": error_msg}), 400


@research_bp.route('/abstracts/<abstract_id>', methods=['PUT'])
@jwt_required()
def update_abstract(abstract_id):
    """Update a research abstract."""
    actor_id, context = _resolve_actor_context("update_abstract")
    abstract = None
    pdf_path: Optional[str] = None

    try:
        abstract = abstract_utils.get_abstract_by_id(
            abstract_id,
            actor_id=actor_id,
            context=context,
        )
        if abstract is None:
            error_msg = f"Resource not found: Abstract with ID {abstract_id} does not exist"
            log_audit_event(
                event_type="abstract.update.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Abstract not found")

        # Check if user is authorized to update this abstract
        user = User.query.get(actor_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="abstract.update.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        # Check authorization - only admin, superadmin, or the creator can update
        if not (
                user.has_role(Role.ADMIN.value) or
                user.has_role(Role.SUPERADMIN.value) or
                abstract.created_by_id == actor_id
        ):
            error_msg = f"Authorization failed: You are not authorized to update abstract ID {abstract_id}"
            log_audit_event(
                event_type="abstract.update.failed",
                user_id=actor_id,
                details={
                    "error": error_msg, 
                    "abstract_id": abstract_id,
                    "user_role": user.role_associations[0].role.value if user.role_associations else "no_role",
                    "abstract_creator_id": abstract.created_by_id
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 403

        if request.is_json:
            payload = request.get_json() or {}
            if not isinstance(payload, dict):
                error_msg = "Request validation failed: Invalid payload provided"
                log_audit_event(
                    event_type="abstract.update.failed",
                    user_id=actor_id,
                    details={"error": error_msg, "abstract_id": abstract_id},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 400
        else:
            data_json = request.form.get("data")
            if not data_json:
                error_msg = "Request validation failed: No data provided in form request"
                log_audit_event(
                    event_type="abstract.update.failed",
                    user_id=actor_id,
                    details={"error": error_msg, "abstract_id": abstract_id},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 400
            try:
                payload = json.loads(data_json)
            except json.JSONDecodeError:
                error_msg = "Request validation failed: Invalid JSON data provided"
                log_audit_event(
                    event_type="abstract.update.failed",
                    user_id=actor_id,
                    details={"error": error_msg, "abstract_id": abstract_id},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 400

            pdf_file = request.files.get("abstract_pdf")
            if pdf_file:
                upload_folder = current_app.config.get("UPLOAD_FOLDER", "uploads")
                import os
                from werkzeug.utils import secure_filename

                os.makedirs(upload_folder, exist_ok=True)
                filename = secure_filename(pdf_file.filename)
                if not filename:
                    error_msg = "File validation failed: Invalid filename provided"
                    log_audit_event(
                        event_type="abstract.update.failed",
                        user_id=actor_id,
                        details={"error": error_msg, "abstract_id": abstract_id, "filename": pdf_file.filename},
                        ip_address=request.remote_addr
                    )
                    return jsonify({"error": error_msg}), 400
                unique_filename = f"{uuid.uuid4().hex}_{filename}"
                file_path = os.path.join(upload_folder, unique_filename)
                pdf_file.save(file_path)
                current_app.logger.info("PDF file saved to: %s", file_path)
                pdf_path = file_path.replace("app/", "", 1)

        authors_data = payload.pop("authors", [])
        if pdf_path:
            payload["pdf_path"] = pdf_path

        abstract = abstract_utils.update_abstract(
            abstract,
            commit=False,
            actor_id=actor_id,
            context=context,
            **payload,
        )

        prev_assocs = AbstractAuthors.query.filter_by(abstract_id=abstract.id).all()
        for assoc in prev_assocs:
            author = Author.query.get(assoc.author_id)
            db.session.delete(assoc)
            if author:
                other_assoc = AbstractAuthors.query.filter(
                    AbstractAuthors.author_id == author.id,
                    AbstractAuthors.abstract_id != abstract.id,
                ).first()
                if not other_assoc:
                    author_utils.delete_author(
                        author,
                        commit=False,
                        actor_id=actor_id,
                        context={**context, "operation": "purge_author"},
                    )

        if authors_data:
            for index, author_data in enumerate(authors_data):
                author = author_utils.create_author(
                    commit=False,
                    actor_id=actor_id,
                    context={**context, "author_index": index},
                    name=author_data.get("name", ""),
                    email=author_data.get("email"),
                    affiliation=author_data.get("affiliation"),
                    is_presenter=author_data.get("is_presenter", False),
                    is_corresponding=author_data.get("is_corresponding", False),
                )
                db.session.flush()
                db.session.add(
                    AbstractAuthors(
                        abstract_id=abstract.id,
                        author_id=author.id,
                        author_order=index,
                    )
                )

        db.session.commit()
        
        # Log successful update
        log_audit_event(
            event_type="abstract.update.success",
            user_id=actor_id,
            details={
                "abstract_id": abstract_id,
                "updated_fields": list(payload.keys()),
                "title": abstract.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(abstract_schema.dump(abstract)), 200
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Error updating abstract")
        error_msg = f"System error occurred while updating abstract: {str(exc)}"
        log_audit_event(
            event_type="abstract.update.failed",
            user_id=actor_id,
            details={"error": error_msg, "abstract_id": abstract_id if abstract else abstract_id, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/abstracts/<abstract_id>/pdf', methods=['GET'])
@jwt_required()
def get_abstract_pdf(abstract_id):
    current_app.logger.info("Fetching PDF for abstract ID: %s", abstract_id)
    actor_id, context = _resolve_actor_context("get_abstract_pdf")
    abstract = None
    try:
        abstract = abstract_utils.get_abstract_by_id(
            abstract_id,
            actor_id=actor_id,
            context=context,
        )
        current_app.logger.info("Abstract fetched: %s", "found" if abstract else "not found")
        if abstract is None:
            error_msg = f"Resource not found: Abstract with ID {abstract_id} does not exist"
            log_audit_event(
                event_type="abstract.pdf.access.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Abstract not found.")

        current_app.logger.info("Checking PDF path for abstract ID: %s", abstract_id)

        if not abstract.pdf_path:
            error_msg = f"File not found: No PDF uploaded for abstract ID {abstract_id}"
            log_audit_event(
                event_type="abstract.pdf.access.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id},
                ip_address=request.remote_addr
            )
            abort(404, description="No PDF uploaded for this abstract.")

        # Log successful access
        log_audit_event(
            event_type="abstract.pdf.access.success",
            user_id=actor_id,
            details={
                "abstract_id": abstract_id,
                "file_path": abstract.pdf_path
            },
            ip_address=request.remote_addr
        )
        
        current_app.logger.info("PDF served successfully")
        return send_file(abstract.pdf_path, mimetype='application/pdf', as_attachment=False)
    except FileNotFoundError:
        error_msg = f"File access error: PDF file not found at path {abstract.pdf_path if 'abstract' in locals() else 'unknown'}"
        log_audit_event(
            event_type="abstract.pdf.access.failed",
            user_id=actor_id,
            details={"error": error_msg, "abstract_id": abstract_id},
            ip_address=request.remote_addr
        )
        abort(404, description="PDF file not found.")
    except Exception as exc:
        current_app.logger.exception("Error sending PDF file")
        error_msg = f"System error occurred while accessing PDF: {str(exc)}"
        log_audit_event(
            event_type="abstract.pdf.access.failed",
            user_id=actor_id,
            details={"error": error_msg, "abstract_id": abstract_id},
            ip_address=request.remote_addr
        )
        abort(500, description="Internal server error.")


@research_bp.route('/abstracts', methods=['GET'])
@jwt_required()
def get_abstracts():
    """Get all research abstracts with filtering and pagination support."""
    actor_id, context = _resolve_actor_context("get_abstracts")
    try:
        q = request.args.get('q', '').strip()
        verifiers = request.args.get('verifiers', '').strip()
        page = max(1, int(request.args.get('page', 1)))
        status = request.args.get('status', '').strip().upper()
        page_size = min(int(request.args.get('page_size', 20)), 100)
        sort_by = request.args.get('sort', 'id')
        sort_dir = request.args.get('dir', 'desc').lower()
        verifier_filter = request.args.get('verifier', '').strip().lower() == 'true'
        # Add review_phase filter
        review_phase = request.args.get('review_phase', '').strip()

        # Get current user for permissions
        user = User.query.get(actor_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="abstract.list.failed",
                user_id=actor_id,
                details={"error": error_msg},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        filters = []
        if q:
            q_int = int(q) if q.isdigit() else None
            filters.append(
                or_(
                    Abstracts.title.ilike(f'%{q}%'),
                    Abstracts.content.ilike(f'%{q}%'),
                    Abstracts.abstract_number == q_int,
                    Abstracts.category.has(Category.name.ilike(f'%{q}%')),

                )
            )

        if status in {'PENDING', 'UNDER_REVIEW', 'ACCEPTED', 'REJECTED'}:
            status_value = Status[status]
            filters.append(Abstracts.status == status_value)
        elif status:  # Invalid status provided
            error_msg = f"Validation failed: Invalid status '{status}'. Valid statuses are PENDING, UNDER_REVIEW, ACCEPTED, REJECTED"
            log_audit_event(
                event_type="abstract.list.failed",
                user_id=actor_id,
                details={"error": error_msg, "invalid_status": status},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400

        # Add review phase filter
        if review_phase:
            try:
                phase_int = int(review_phase)
                filters.append(Abstracts.review_phase == phase_int)
            except ValueError:
                error_msg = f"Validation failed: Invalid review phase '{review_phase}'. Review phase must be a number."
                log_audit_event(
                    event_type="abstract.list.failed",
                    user_id=actor_id,
                    details={"error": error_msg, "invalid_review_phase": review_phase},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 400

        privileged = any(
            user.has_role(role)
            for role in (
                Role.ADMIN.value,
                Role.SUPERADMIN.value,
                Role.VERIFIER.value,
                Role.COORDINATOR.value,
            )
        )
        if not privileged:
            filters.append(Abstracts.created_by_id == actor_id)

        if verifiers:
            if verifiers.lower() == 'yes':
                filters.append(Abstracts.verifiers.any())
            elif verifiers.lower() == 'no':
                filters.append(~Abstracts.verifiers.any())
            else:
                error_msg = f"Validation failed: Invalid verifiers parameter '{verifiers}'. Valid values are 'yes' or 'no'"
                log_audit_event(
                    event_type="abstract.list.failed",
                    user_id=actor_id,
                    details={"error": error_msg, "invalid_verifiers_param": verifiers},
                    ip_address=request.remote_addr
                )
                return jsonify({"error": error_msg}), 400

        if verifier_filter and actor_id:
            filters.append(Abstracts.verifiers.any(User.id == actor_id))

        current_app.logger.info(f"Sorting by {sort_by} in {sort_dir} order")

        if sort_by == 'title':
            order_by = Abstracts.title.asc() if sort_dir == 'asc' else Abstracts.title.desc()
        elif sort_by == 'created_at':
            order_by = Abstracts.created_at.asc() if sort_dir == 'asc' else Abstracts.created_at.desc()
        elif sort_by == 'id':
            order_by = Abstracts.id.asc() if sort_dir == 'asc' else Abstracts.id.desc()
        elif sort_by == 'review_phase':  # Add sorting by review phase
            order_by = Abstracts.review_phase.asc() if sort_dir == 'asc' else Abstracts.review_phase.desc()
        else: # invalid sort field
            error_msg = f"Validation failed: Invalid sort field '{sort_by}'. Valid fields are 'id', 'title', 'created_at', 'review_phase'"
            log_audit_event(
                event_type="abstract.list.failed",
                user_id=actor_id,
                details={"error": error_msg, "invalid_sort_by": sort_by},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400

        abstracts_all = list(
            abstract_utils.list_abstracts(
                filters=filters,
                order_by=order_by,
                actor_id=actor_id,
                context={**context, "sort_by": sort_by, "sort_dir": sort_dir},
            )
        )

        total = len(abstracts_all)
        start = (page - 1) * page_size
        end = start + page_size
        paginated = abstracts_all[start:end]

        abstracts_data = []
        for abstract in paginated:
            abstract_dict = abstract_schema.dump(abstract)
            abstract_dict['verifiers_count'] = len(abstract.verifiers or [])
            abstract_dict['review_phase'] = abstract.review_phase  # Include review phase in response
            abstracts_data.append(abstract_dict)

        response = {
            'items': abstracts_data,
            'total': total,
            'page': page,
            'pages': (total + page_size - 1) // page_size,
            'page_size': page_size
        }

        # Log successful retrieval
        log_audit_event(
            event_type="abstract.list.success",
            user_id=actor_id,
            details={
                "filters_applied": bool(filters),
                "search_query": q if q else None,
                "status_filter": status if status else None,
                "review_phase_filter": review_phase if review_phase else None,
                "verifier_filter": verifier_filter,
                "results_count": len(abstracts_data),
                "total_count": total,
                "page": page,
                "page_size": page_size
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(response), 200
    except Exception as exc:
        current_app.logger.exception("Error listing abstracts with parameters")
        error_msg = f"System error occurred while retrieving abstracts: {str(exc)}"
        log_audit_event(
            event_type="abstract.list.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/abstracts/<abstract_id>', methods=['GET'])
@jwt_required()
def get_abstract(abstract_id):
    """Get a specific research abstract."""
    actor_id, context = _resolve_actor_context("get_abstract")
    abstract = None
    try:
        abstract = abstract_utils.get_abstract_by_id(
            abstract_id,
            actor_id=actor_id,
            context=context,
        )
        if abstract is None:
            error_msg = f"Resource not found: Abstract with ID {abstract_id} does not exist"
            log_audit_event(
                event_type="abstract.get.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Abstract not found.")

        data = abstract_schema.dump(abstract)
        if abstract.pdf_path:
            data['pdf_url'] = f"/api/v1/research/abstracts/{abstract_id}/pdf"
        # Include review phase in the response
        data['review_phase'] = abstract.review_phase

        # Log successful retrieval
        log_audit_event(
            event_type="abstract.get.success",
            user_id=actor_id,
            details={
                "abstract_id": abstract_id,
                "title": abstract.title,
                "has_pdf": bool(abstract.pdf_path),
                "review_phase": abstract.review_phase
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(data), 200
    except Exception as exc:
        current_app.logger.exception("Error retrieving abstract")
        error_msg = f"System error occurred while retrieving abstract: {str(exc)}"
        log_audit_event(
            event_type="abstract.get.failed",
            user_id=actor_id,
            details={"error": error_msg, "abstract_id": abstract_id if abstract else abstract_id, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/abstracts/<abstract_id>', methods=['DELETE'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def delete_abstract(abstract_id):
    """Delete a research abstract."""
    actor_id, context = _resolve_actor_context("delete_abstract")
    abstract = None
    try:
        abstract = abstract_utils.get_abstract_by_id(
            abstract_id,
            actor_id=actor_id,
            context=context,
        )
        if abstract is None:
            error_msg = f"Resource not found: Abstract with ID {abstract_id} does not exist"
            log_audit_event(
                event_type="abstract.delete.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Abstract not found.")
            
        abstract_utils.delete_abstract(
            abstract,
            actor_id=actor_id,
            context=context,
        )
        
        # Log successful deletion
        log_audit_event(
            event_type="abstract.delete.success",
            user_id=actor_id,
            details={
                "abstract_id": abstract_id,
                "title": abstract.title,
                "deleted_by": actor_id
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Abstract deleted"}), 200
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Error deleting abstract")
        error_msg = f"System error occurred while deleting abstract: {str(exc)}"
        log_audit_event(
            event_type="abstract.delete.failed",
            user_id=actor_id,
            details={"error": error_msg, "abstract_id": abstract_id if abstract else abstract_id, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/abstracts/<abstract_id>/submit', methods=['POST'])
@jwt_required()
def submit_abstract(abstract_id):
    """Submit an abstract for review."""
    actor_id, context = _resolve_actor_context("submit_abstract")
    abstract = None
    try:
        abstract = abstract_utils.get_abstract_by_id(
            abstract_id,
            actor_id=actor_id,
            context=context,
        )
        if abstract is None:
            error_msg = f"Resource not found: Abstract with ID {abstract_id} does not exist"
            log_audit_event(
                event_type="abstract.submit.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Abstract not found.")

        # Check if user is authorized to submit this abstract
        user = User.query.get(actor_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="abstract.submit.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404
            
        # Only the creator of the abstract or admin can submit it
        if not (
                abstract.created_by_id == actor_id or
                user.has_role(Role.ADMIN.value) or
                user.has_role(Role.SUPERADMIN.value)
        ):
            error_msg = f"Authorization failed: You are not authorized to submit abstract ID {abstract_id} for review"
            log_audit_event(
                event_type="abstract.submit.failed",
                user_id=actor_id,
                details={
                    "error": error_msg, 
                    "abstract_id": abstract_id,
                    "user_role": user.role_associations[0].role.value if user.role_associations else "no_role",
                    "abstract_creator_id": abstract.created_by_id
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 403

        # Only allow submission if the abstract is in PENDING status and phase 1
        if abstract.status != Status.PENDING or abstract.review_phase != 1:
            error_msg = f"Cannot submit abstract: Abstract must be in PENDING status and phase 1 to be submitted for review. Current status: {abstract.status.name}, Current phase: {abstract.review_phase}"
            log_audit_event(
                event_type="abstract.submit.failed",
                user_id=actor_id,
                details={
                    "error": error_msg,
                    "abstract_id": abstract_id,
                    "current_status": abstract.status.name,
                    "current_phase": abstract.review_phase
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400

        abstract_utils.update_abstract(
            abstract,
            actor_id=actor_id,
            context=context,
            status=Status.UNDER_REVIEW,
        )
        
        # Log successful submission
        log_audit_event(
            event_type="abstract.submit.success",
            user_id=actor_id,
            details={
                "abstract_id": abstract_id,
                "new_status": "UNDER_REVIEW",
                "title": abstract.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Abstract submitted for review"}), 200
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Error submitting abstract")
        error_msg = f"System error occurred while submitting abstract: {str(exc)}"
        log_audit_event(
            event_type="abstract.submit.failed",
            user_id=actor_id,
            details={"error": error_msg, "abstract_id": abstract_id if abstract else abstract_id, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400
    

@research_bp.route('/abstracts/status', methods=['GET'])
@jwt_required()
def get_abstract_submission_status():
    """Get submission status of abstracts for the current user."""
    actor_id, context = _resolve_actor_context("get_abstract_submission_status")
    user = None
    try:
        user = User.query.get(actor_id)
        if user is None:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="abstract.status.get.failed",
                user_id=actor_id,
                details={"error": error_msg},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        if user.has_role(Role.ADMIN.value) or user.has_role(Role.SUPERADMIN.value):
            abstracts = abstract_utils.list_abstracts(
                actor_id=actor_id,
                context=context,
            )
        elif user.has_role(Role.VERIFIER.value):
            current_app.logger.info("Fetching abstracts for verifier user ID: %s", actor_id)
            abstracts = abstract_utils.list_abstracts(
                filters=[Abstracts.verifiers.any(User.id == actor_id)],
                actor_id=actor_id,
                context={**context, "scope": "verifier"},
            )
        elif user.has_role(Role.COORDINATOR.value):
            current_app.logger.info("Fetching abstracts for coordinator user ID: %s", actor_id)
            abstracts = abstract_utils.list_abstracts(
                filters=[Abstracts.coordinators.any(User.id == actor_id)],
                actor_id=actor_id,
                context={**context, "scope": "coordinator"},
            )
        else:
            abstracts = abstract_utils.list_abstracts(
                filters=[Abstracts.created_by_id == actor_id],
                actor_id=actor_id,
                context={**context, "scope": "owner"},
            )

        pending_abstracts = sum(1 for abstract in abstracts if abstract.status == Status.PENDING)
        under_review_abstracts = sum(1 for abstract in abstracts if abstract.status == Status.UNDER_REVIEW)
        accepted_abstracts = sum(1 for abstract in abstracts if abstract.status == Status.ACCEPTED)
        rejected_abstracts = sum(1 for abstract in abstracts if abstract.status == Status.REJECTED)

        payload = {
            "pending": pending_abstracts,
            "under_review": under_review_abstracts,
            "accepted": accepted_abstracts,
            "rejected": rejected_abstracts,
        }

        # Log successful retrieval
        log_audit_event(
            event_type="abstract.status.get.success",
            user_id=actor_id,
            details={
                "user_role": user.role_associations[0].role.value if user.role_associations else "no_role",
                "pending_count": pending_abstracts,
                "under_review_count": under_review_abstracts,
                "accepted_count": accepted_abstracts,
                "rejected_count": rejected_abstracts
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(payload), 200
    except Exception as exc:
        current_app.logger.exception("Error retrieving abstract submission status")
        error_msg = f"System error occurred while retrieving abstract status: {str(exc)}"
        log_audit_event(
            event_type="abstract.status.get.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


# Verifier Management Routes

@research_bp.route('/abstracts/<abstract_id>/verifiers/<user_id>', methods=['POST'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def assign_verifier_to_abstract(abstract_id, user_id):
    """Assign a verifier to an abstract."""
    actor_id, context = _resolve_actor_context("assign_verifier_to_abstract")
    abstract = None
    user = None
    try:
        abstract = abstract_utils.get_abstract_by_id(
            abstract_id,
            actor_id=actor_id,
            context=context,
        )
        if abstract is None:
            error_msg = f"Resource not found: Abstract with ID {abstract_id} does not exist"
            log_audit_event(
                event_type="abstract.verifier.assign.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Abstract not found.")

        user = User.query.get(user_id)
        if user is None:
            error_msg = f"Resource not found: User with ID {user_id} does not exist"
            log_audit_event(
                event_type="abstract.verifier.assign.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            abort(404, description="User not found.")

        if not user.has_role(Role.VERIFIER.value):
            error_msg = f"Validation failed: User with ID {user_id} is not a verifier"
            log_audit_event(
                event_type="abstract.verifier.assign.failed",
                user_id=actor_id,
                details={
                    "error": error_msg, 
                    "abstract_id": abstract_id, 
                    "verifier_id": user_id,
                    "user_role": user.role_associations[0].role.value if user.role_associations else "no_role"
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400

        if any(verifier.id == user.id for verifier in abstract.verifiers):
            error_msg = "Verifier already assigned to this abstract"
            log_audit_event(
                event_type="abstract.verifier.assign.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            return jsonify({"message": error_msg}), 200

        abstract_utils.assign_verifier(
            abstract,
            user,
            actor_id=actor_id,
            context={**context, "verifier_id": str(user_id)},
        )

        # Log successful assignment
        log_audit_event(
            event_type="abstract.verifier.assign.success",
            user_id=actor_id,
            details={
                "abstract_id": abstract_id,
                "verifier_id": user_id,
                "verifier_username": user.username,
                "title": abstract.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Verifier assigned successfully"}), 201
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Error assigning verifier to abstract")
        error_msg = f"System error occurred while assigning verifier to abstract: {str(exc)}"
        log_audit_event(
            event_type="abstract.verifier.assign.failed",
            user_id=actor_id,
            details={"error": error_msg, "abstract_id": abstract_id if abstract else abstract_id, "verifier_id": user_id if user else user_id, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/abstracts/<abstract_id>/verifiers/<user_id>', methods=['DELETE'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def unassign_verifier_from_abstract(abstract_id, user_id):
    """Unassign a verifier from an abstract."""
    actor_id, context = _resolve_actor_context("unassign_verifier_from_abstract")
    abstract = None
    user = None
    try:
        abstract = abstract_utils.get_abstract_by_id(
            abstract_id,
            actor_id=actor_id,
            context=context,
        )
        if abstract is None:
            error_msg = f"Resource not found: Abstract with ID {abstract_id} does not exist"
            log_audit_event(
                event_type="abstract.verifier.unassign.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Abstract not found.")

        user = User.query.get(user_id)
        if user is None:
            error_msg = f"Resource not found: User with ID {user_id} does not exist"
            log_audit_event(
                event_type="abstract.verifier.unassign.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            abort(404, description="User not found.")

        if not any(verifier.id == user.id for verifier in abstract.verifiers):
            error_msg = "Verifier not assigned to this abstract"
            log_audit_event(
                event_type="abstract.verifier.unassign.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        abstract_utils.remove_verifier(
            abstract,
            user,
            actor_id=actor_id,
            context={**context, "verifier_id": str(user_id)},
        )

        # Log successful unassignment
        log_audit_event(
            event_type="abstract.verifier.unassign.success",
            user_id=actor_id,
            details={
                "abstract_id": abstract_id,
                "verifier_id": user_id,
                "verifier_username": user.username,
                "title": abstract.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Verifier unassigned successfully"}), 200
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Error unassigning verifier from abstract")
        error_msg = f"System error occurred while unassigning verifier from abstract: {str(exc)}"
        log_audit_event(
            event_type="abstract.verifier.unassign.failed",
            user_id=actor_id,
            details={"error": error_msg, "abstract_id": abstract_id if abstract else abstract_id, "verifier_id": user_id if user else user_id, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/abstracts/<abstract_id>/verifiers', methods=['GET'])
@jwt_required()
def get_verifiers_for_abstract(abstract_id):
    """Get all verifiers assigned to an abstract."""
    actor_id, context = _resolve_actor_context("get_verifiers_for_abstract")
    abstract = None
    try:
        abstract = abstract_utils.get_abstract_by_id(
            abstract_id,
            actor_id=actor_id,
            context=context,
        )
        if abstract is None:
            error_msg = f"Resource not found: Abstract with ID {abstract_id} does not exist"
            log_audit_event(
                event_type="abstract.verifiers.get.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Abstract not found.")

        verifiers_data = [
            {
                'id': str(verifier.id),
                'username': verifier.username,
                'email': verifier.email,
                'employee_id': verifier.employee_id,
            }
            for verifier in abstract.verifiers
        ]

        # Log successful retrieval
        log_audit_event(
            event_type="abstract.verifiers.get.success",
            user_id=actor_id,
            details={
                "abstract_id": abstract_id,
                "verifiers_count": len(verifiers_data),
                "verifiers": [v['username'] for v in verifiers_data]
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(verifiers_data), 200
    except Exception as exc:
        current_app.logger.exception("Error getting verifiers for abstract")
        error_msg = f"System error occurred while retrieving verifiers for abstract: {str(exc)}"
        log_audit_event(
            event_type="abstract.verifiers.get.failed",
            user_id=actor_id,
            details={"error": error_msg, "abstract_id": abstract_id if abstract else abstract_id, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/verifiers/<user_id>/abstracts', methods=['GET'])
@jwt_required()
def get_abstracts_for_verifier(user_id):
    """Get all abstracts assigned to a verifier."""
    actor_id, context = _resolve_actor_context("get_abstracts_for_verifier")
    user = None
    try:
        user = User.query.get(user_id)
        if user is None:
            error_msg = f"Resource not found: User with ID {user_id} does not exist"
            log_audit_event(
                event_type="verifier.abstracts.get.failed",
                user_id=actor_id,
                details={"error": error_msg, "verifier_id": user_id},
                ip_address=request.remote_addr
            )
            abort(404, description="User not found.")

        abstracts = abstract_utils.list_abstracts(
            filters=[Abstracts.verifiers.any(User.id == user_id)],
            actor_id=actor_id,
            context={**context, "target_verifier": str(user_id)},
        )

        # Log successful retrieval
        log_audit_event(
            event_type="verifier.abstracts.get.success",
            user_id=actor_id,
            details={
                "verifier_id": user_id,
                "abstracts_count": len(abstracts),
                "abstracts": [abstract.title for abstract in abstracts]
            },
            ip_address=request.remote_addr
        )
        
        return jsonify(abstracts_schema.dump(abstracts)), 200
    except Exception as exc:
        current_app.logger.exception("Error getting abstracts for verifier")
        error_msg = f"System error occurred while retrieving abstracts for verifier: {str(exc)}"
        log_audit_event(
            event_type="verifier.abstracts.get.failed",
            user_id=actor_id,
            details={"error": error_msg, "verifier_id": user_id if user else user_id, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/abstracts/bulk-assign-verifiers', methods=['POST'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def bulk_assign_verifiers():
    """Bulk assign verifiers to multiple abstracts."""
    actor_id, context = _resolve_actor_context("bulk_assign_verifiers")
    try:
        data = request.get_json() or {}
        abstract_ids = data.get('abstract_ids')
        user_ids = data.get('user_ids')

        # Validate input
        if not abstract_ids or not user_ids:
            error_msg = "Request validation failed: Missing required fields 'abstract_ids' or 'user_ids' in request body"
            log_audit_event(
                event_type="abstract.verifiers.bulk_assign.failed",
                user_id=actor_id,
                details={"error": error_msg, "provided_fields": list(data.keys()) if data else []},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400

        abstract_ids = [str(abstract_id) for abstract_id in abstract_ids]
        user_ids = [str(user_id) for user_id in user_ids]

        abstracts = abstract_utils.list_abstracts(
            filters=[Abstracts.id.in_(abstract_ids)],
            actor_id=actor_id,
            context={**context, "operation": "bulk_assign"},
        )
        abstracts_by_id = {str(abstract.id): abstract for abstract in abstracts}
        missing_abstracts = [abstract_id for abstract_id in abstract_ids if abstract_id not in abstracts_by_id]
        if missing_abstracts:
            error_msg = f"Resource validation failed: Abstracts with IDs {missing_abstracts} do not exist"
            log_audit_event(
                event_type="abstract.verifiers.bulk_assign.failed",
                user_id=actor_id,
                details={"error": error_msg, "missing_abstracts": missing_abstracts},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        users = User.query.filter(User.id.in_(user_ids)).all()
        if len(users) != len(set(user_ids)):
            found_ids = {str(user.id) for user in users}
            missing_ids = [uid for uid in user_ids if uid not in found_ids]
            error_msg = f"Resource validation failed: Users with IDs {missing_ids} do not exist"
            log_audit_event(
                event_type="abstract.verifiers.bulk_assign.failed",
                user_id=actor_id,
                details={"error": error_msg, "missing_users": missing_ids},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        non_verifiers = [user for user in users if not user.has_role(Role.VERIFIER.value)]
        if non_verifiers:
            non_verifier_ids = [str(user.id) for user in non_verifiers]
            error_msg = f"Validation failed: Users with IDs {non_verifier_ids} are not verifiers"
            log_audit_event(
                event_type="abstract.verifiers.bulk_assign.failed",
                user_id=actor_id,
                details={"error": error_msg, "non_verifier_ids": non_verifier_ids},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400

        assignments_created = 0
        for abstract_id in abstract_ids:
            abstract = abstracts_by_id[str(abstract_id)]
            for user in users:
                if any(verifier.id == user.id for verifier in abstract.verifiers):
                    continue
                abstract_utils.assign_verifier(
                    abstract,
                    user,
                    actor_id=actor_id,
                    context={**context, "verifier_id": str(user.id), "abstract_id": str(abstract.id)},
                )
                assignments_created += 1

        # Log successful bulk assignment
        log_audit_event(
            event_type="abstract.verifiers.bulk_assign.success",
            user_id=actor_id,
            details={
                "abstract_ids": abstract_ids,
                "user_ids": user_ids,
                "assignments_created": assignments_created,
                "total_possible_assignments": len(abstract_ids) * len(users)
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({
            "message": f"Successfully created {assignments_created} assignments",
            "assignments_created": assignments_created
        }), 201
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Error in bulk assigning verifiers")
        error_msg = f"System error occurred during bulk assignment of verifiers: {str(exc)}"
        log_audit_event(
            event_type="abstract.verifiers.bulk_assign.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/abstracts/bulk-unassign-verifiers', methods=['POST'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def bulk_unassign_verifiers():
    """Bulk unassign verifiers from multiple abstracts."""
    actor_id, context = _resolve_actor_context("bulk_unassign_verifiers")
    try:
        data = request.get_json() or {}

        abstract_ids = data.get('abstract_ids')
        user_ids = data.get('user_ids')

        # Validate input
        if not abstract_ids or not user_ids:
            error_msg = "Request validation failed: Missing required fields 'abstract_ids' or 'user_ids' in request body"
            log_audit_event(
                event_type="abstract.verifiers.bulk_unassign.failed",
                user_id=actor_id,
                details={"error": error_msg, "provided_fields": list(data.keys()) if data else []},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400

        abstract_ids = [str(abstract_id) for abstract_id in abstract_ids]
        user_ids = [str(user_id) for user_id in user_ids]

        abstracts = abstract_utils.list_abstracts(
            filters=[Abstracts.id.in_(abstract_ids)],
            actor_id=actor_id,
            context={**context, "operation": "bulk_unassign"},
        )
        abstracts_by_id = {str(abstract.id): abstract for abstract in abstracts}
        missing_abstracts = [abstract_id for abstract_id in abstract_ids if abstract_id not in abstracts_by_id]
        if missing_abstracts:
            error_msg = f"Resource validation failed: Abstracts with IDs {missing_abstracts} do not exist"
            log_audit_event(
                event_type="abstract.verifiers.bulk_unassign.failed",
                user_id=actor_id,
                details={"error": error_msg, "missing_abstracts": missing_abstracts},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        users = User.query.filter(User.id.in_(user_ids)).all()
        if len(users) != len(set(user_ids)):
            found_ids = {str(user.id) for user in users}
            missing_ids = [uid for uid in user_ids if uid not in found_ids]
            error_msg = f"Resource validation failed: Users with IDs {missing_ids} do not exist"
            log_audit_event(
                event_type="abstract.verifiers.bulk_unassign.failed",
                user_id=actor_id,
                details={"error": error_msg, "missing_users": missing_ids},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        assignments_deleted = 0
        for abstract_id in abstract_ids:
            abstract = abstracts_by_id[str(abstract_id)]
            for user in users:
                if any(verifier.id == user.id for verifier in abstract.verifiers):
                    abstract_utils.remove_verifier(
                        abstract,
                        user,
                        actor_id=actor_id,
                        context={**context, "verifier_id": str(user.id), "abstract_id": str(abstract.id)},
                    )
                    assignments_deleted += 1

        # Explicitly commit the transaction
        if not safe_commit():
            error_msg = "Database commit failed during bulk unassignment"
            log_audit_event(
                event_type="abstract.verifiers.bulk_unassign.failed",
                user_id=actor_id,
                details={"error": error_msg},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 500

        # Log successful bulk unassignment
        log_audit_event(
            event_type="abstract.verifiers.bulk_unassign.success",
            user_id=actor_id,
            details={
                "abstract_ids": abstract_ids,
                "user_ids": user_ids,
                "assignments_deleted": assignments_deleted
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({
            "message": f"Successfully deleted {assignments_deleted} assignments",
            "assignments_deleted": assignments_deleted
        }), 200
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Error in bulk unassigning verifiers")
        error_msg = f"System error occurred during bulk unassignment of verifiers: {str(exc)}"
        log_audit_event(
            event_type="abstract.verifiers.bulk_unassign.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


# New endpoint to accept an abstract with grades
@research_bp.route('/abstracts/<abstract_id>/accept', methods=['POST'])
@jwt_required()
@require_roles(Role.VERIFIER,Role.COORDINATOR.value, Role.ADMIN.value, Role.SUPERADMIN.value)
def accept_abstract(abstract_id):
    """Accept an abstract after review."""
    actor_id, context = _resolve_actor_context("accept_abstract")
    abstract = None
    try:
        abstract = abstract_utils.get_abstract_by_id(
            abstract_id,
            actor_id=actor_id,
            context=context,
        )
        if abstract is None:
            error_msg = f"Resource not found: Abstract with ID {abstract_id} does not exist"
            log_audit_event(
                event_type="abstract.accept.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Abstract not found.")

        # Check if user is authorized to accept this abstract
        user = User.query.get(actor_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="abstract.accept.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        # Allow if user is admin, superadmin, or a coordinator for this abstract
        privileged = any(
            user.has_role(role)
            for role in (Role.VERIFIER.value,Role.COORDINATOR.value, Role.ADMIN.value, Role.SUPERADMIN.value)
        )
        
        if not privileged and not any(coordinator.id == actor_id for coordinator in abstract.coordinators):
            error_msg = f"Authorization failed: You are not authorized to accept abstract ID {abstract_id}"
            log_audit_event(
                event_type="abstract.accept.failed",
                user_id=actor_id,
                details={
                    "error": error_msg, 
                    "abstract_id": abstract_id,
                    "user_id": actor_id,
                    "user_role": user.role_associations[0].role.value if user.role_associations else "no_role"
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 403

        # Check if all required grades have been submitted for the current phase
        if not abstract_utils.can_advance_to_next_phase(abstract):
            error_msg = f"Cannot accept abstract: Not all required grades have been submitted for phase {abstract.review_phase}"
            log_audit_event(
                event_type="abstract.accept.failed",
                user_id=actor_id,
                details={
                    "error": error_msg, 
                    "abstract_id": abstract_id,
                    "current_phase": abstract.review_phase
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 400

        # Update the status to ACCEPTED
        abstract_utils.update_abstract(
            abstract,
            actor_id=actor_id,
            context=context,
            status=Status.ACCEPTED,
        )
        
        # Log successful acceptance
        log_audit_event(
            event_type="abstract.accept.success",
            user_id=actor_id,
            details={
                "abstract_id": abstract_id,
                "new_status": "ACCEPTED",
                "title": abstract.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Abstract accepted successfully"}), 200
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Error accepting abstract")
        error_msg = f"System error occurred while accepting abstract: {str(exc)}"
        log_audit_event(
            event_type="abstract.accept.failed",
            user_id=actor_id,
            details={"error": error_msg, "abstract_id": abstract_id if abstract else abstract_id, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


# New endpoint to reject an abstract
@research_bp.route('/abstracts/<abstract_id>/reject', methods=['POST'])
@jwt_required()
@require_roles(Role.COORDINATOR.value, Role.ADMIN.value, Role.SUPERADMIN.value)
def reject_abstract(abstract_id):
    """Reject an abstract after review."""
    actor_id, context = _resolve_actor_context("reject_abstract")
    abstract = None
    try:
        abstract = abstract_utils.get_abstract_by_id(
            abstract_id,
            actor_id=actor_id,
            context=context,
        )
        if abstract is None:
            error_msg = f"Resource not found: Abstract with ID {abstract_id} does not exist"
            log_audit_event(
                event_type="abstract.reject.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id},
                ip_address=request.remote_addr
            )
            abort(404, description="Abstract not found.")

        # Check if user is authorized to reject this abstract
        user = User.query.get(actor_id)
        if not user:
            error_msg = "Authentication failed: User not found"
            log_audit_event(
                event_type="abstract.reject.failed",
                user_id=actor_id,
                details={"error": error_msg, "abstract_id": abstract_id},
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 404

        # Allow if user is admin, superadmin, or a coordinator for this abstract
        privileged = any(
            user.has_role(role)
            for role in (Role.ADMIN.value, Role.SUPERADMIN.value)
        )
        
        if not privileged and not any(coordinator.id == actor_id for coordinator in abstract.coordinators):
            error_msg = f"Authorization failed: You are not authorized to reject abstract ID {abstract_id}"
            log_audit_event(
                event_type="abstract.reject.failed",
                user_id=actor_id,
                details={
                    "error": error_msg, 
                    "abstract_id": abstract_id,
                    "user_id": actor_id,
                    "user_role": user.role_associations[0].role.value if user.role_associations else "no_role"
                },
                ip_address=request.remote_addr
            )
            return jsonify({"error": error_msg}), 403

        # Update the status to REJECTED
        abstract_utils.update_abstract(
            abstract,
            actor_id=actor_id,
            context=context,
            status=Status.REJECTED,
        )
        
        # Log successful rejection
        log_audit_event(
            event_type="abstract.reject.success",
            user_id=actor_id,
            details={
                "abstract_id": abstract_id,
                "new_status": "REJECTED",
                "title": abstract.title
            },
            ip_address=request.remote_addr
        )
        
        return jsonify({"message": "Abstract rejected successfully"}), 200
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("Error rejecting abstract")
        error_msg = f"System error occurred while rejecting abstract: {str(exc)}"
        log_audit_event(
            event_type="abstract.reject.failed",
            user_id=actor_id,
            details={"error": error_msg, "abstract_id": abstract_id if abstract else abstract_id, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


def create_abstracts_excel(abstracts):
    """Create an Excel workbook with all abstracts data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Abstracts Master Sheet"
    
    # Define headers with more detailed information from the models
    headers = [
        "ID", "Abstract Number", "Title", "Category", "Content", "Status", 
        "Created By", "Created At", "Updated At", "PDF Path", 
        "Review Phase", "Cycle Name", "Cycle Start Date", "Cycle End Date", 
        "Authors", "Author Emails", "Affiliations", "Presenter", "Corresponding"
    ]
    
    # Add headers to worksheet
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add abstract data to worksheet
    for row_num, abstract in enumerate(abstracts, 2):
        # Get author information
        authors_list = [author.name for author in abstract.authors] if abstract.authors else []
        author_emails_list = [author.email for author in abstract.authors] if abstract.authors else []
        affiliations_list = [author.affiliation for author in abstract.authors] if abstract.authors else []
        is_presenter_list = [author.name for author in abstract.authors if author.is_presenter] if abstract.authors else []
        is_corresponding_list = [author.name for author in abstract.authors if author.is_corresponding] if abstract.authors else []
        

        # Get creator name
        creator_name = abstract.created_by.username if abstract.created_by_id else "N/A"
        
        # Get cycle information
        cycle_name = abstract.cycle.name if abstract.cycle else "N/A"
        cycle_start_date = abstract.cycle.start_date.isoformat() if abstract.cycle and abstract.cycle.start_date else "N/A"
        cycle_end_date = abstract.cycle.end_date.isoformat() if abstract.cycle and abstract.cycle.end_date else "N/A"
        
        # Get category information
        category_name = abstract.category.name if abstract.category else "N/A"
        
        # Add row data
        row_data = [
            str(abstract.id),  # Convert UUID to string
            abstract.abstract_number,
            abstract.title,
            category_name,
            abstract.content[:200] + "..." if len(abstract.content) > 200 else abstract.content,  # Truncate long content
            abstract.status.name if abstract.status else "N/A",
            creator_name,
            abstract.created_at.isoformat() if abstract.created_at else "N/A",
            abstract.updated_at.isoformat() if abstract.updated_at else "N/A",
            abstract.pdf_path or "N/A",
            abstract.review_phase,
            cycle_name,
            cycle_start_date,
            cycle_end_date,
            "; ".join(authors_list),
            "; ".join(str(e) for e in author_emails_list if e),
            "; ".join(str(e) for e in affiliations_list if e),
            "; ".join(str(e) for e in is_presenter_list if e),
            "; ".join(str(e) for e in is_corresponding_list if e)
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Limit max width to 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb


@research_bp.route('/abstracts/export-excel', methods=['GET'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def export_abstracts_excel():
    """Export all abstracts to an Excel file."""
    actor_id, context = _resolve_actor_context("export_abstracts_excel")
    try:
        # Get all abstracts with related data
        abstracts = abstract_utils.list_abstracts(
            actor_id=actor_id,
            context=context
        )
        
        # Create Excel workbook
        wb = create_abstracts_excel(abstracts)
        
        # Save to a temporary file
        from io import BytesIO
        from flask import send_file
        import tempfile
        import os
        
        # Create a temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()  # Close the file so it can be read by send_file
        
        # Log successful export
        log_audit_event(
            event_type="abstract.excel.export.success",
            user_id=actor_id,
            details={
                "exported_count": len(abstracts),
                "file_path": temp_file.name
            },
            ip_address=request.remote_addr
        )
        
        # Send the file to the user
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f"abstracts_master_{uuid.uuid4().hex[:8]}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as exc:
        current_app.logger.exception("Error exporting abstracts to Excel")
        error_msg = f"System error occurred while exporting abstracts to Excel: {str(exc)}"
        log_audit_event(
            event_type="abstract.excel.export.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400


@research_bp.route('/abstracts/export-pdf-zip', methods=['GET'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def export_abstracts_pdf_zip():
    """Export all abstract PDFs in a ZIP file organized by category, including an Excel summary."""
    actor_id, context = _resolve_actor_context("export_abstracts_pdf_zip")
    try:
        import zipfile
        import shutil
        from io import BytesIO
        import tempfile
        
        # Get all abstracts with related data
        abstracts = abstract_utils.list_abstracts(
            actor_id=actor_id,
            context=context
        )
        
        # Base path where PDF files are stored
        base_path = current_app.config.get("UPLOAD_FOLDER", "uploads")
        
        # Create a temporary directory to organize files
        temp_dir = tempfile.mkdtemp()
        
        # Organize abstracts by category
        categories = {}
        for abstract in abstracts:
            if abstract.pdf_path and abstract.category:
                category_name = abstract.category.name
                if category_name not in categories:
                    categories[category_name] = []
                categories[category_name].append(abstract)
        
        # Create subdirectories for each category and copy PDF files
        for category_name, abstract_list in categories.items():
            category_dir = os.path.join(temp_dir, category_name)
            os.makedirs(category_dir, exist_ok=True)
            
            for abstract in abstract_list:
                if abstract.pdf_path:
                    # Construct the full path from the base path
                    full_pdf_path = os.path.join(base_path, os.path.basename(abstract.pdf_path))
                    
                    # Verify the file exists before copying
                    if os.path.exists(full_pdf_path):
                        # Get the original filename
                        original_filename = os.path.basename(full_pdf_path)
                        # Create new filename with abstract ID to avoid conflicts
                        new_filename = f"{abstract.id.hex}_{original_filename}"
                        destination_path = os.path.join(category_dir, new_filename)
                        
                        # Copy the PDF file to the category folder
                        shutil.copy2(full_pdf_path, destination_path)
                    else:
                        current_app.logger.warning(f"PDF file not found at path: {full_pdf_path}")
        
        # Generate Excel file with all abstracts data using the existing function
        excel_wb = create_abstracts_excel(abstracts)
        
        # Save the Excel file to the temporary directory root
        excel_path = os.path.join(temp_dir, f"abstracts_summary_{uuid.uuid4().hex[:8]}.xlsx")
        excel_wb.save(excel_path)
        
        # Create a ZIP file from the organized directory structure
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    # Calculate the archive name (path inside the ZIP)
                    archive_name = os.path.relpath(file_path, temp_dir)
                    zip_file.write(file_path, archive_name)
        
        # Clean up the temporary directory
        shutil.rmtree(temp_dir)
        
        # Prepare the ZIP file for download
        zip_buffer.seek(0)
        
        # Log successful export
        log_audit_event(
            event_type="abstract.pdf.excel.zip.export.success",
            user_id=actor_id,
            details={
                "exported_count": len(abstracts),
                "categories_count": len(categories),
                "categories": list(categories.keys())
            },
            ip_address=request.remote_addr
        )
        
        # Send the ZIP file to the user
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=f"abstracts_pdfs_with_excel_{uuid.uuid4().hex[:8]}.zip",
            mimetype='application/zip'
        )
        
    except Exception as exc:
        current_app.logger.exception("Error exporting abstracts PDFs to ZIP")
        error_msg = f"System error occurred while exporting abstracts PDFs to ZIP: {str(exc)}"
        log_audit_event(
            event_type="abstract.pdf.excel.zip.export.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400
@research_bp.route('/abstracts/export-with-pdfs', methods=['GET'])
@jwt_required()
@require_roles(Role.ADMIN.value, Role.SUPERADMIN.value)
def export_abstracts_with_pdfs():
    """Export all abstracts with PDFs organized by category in a ZIP file, including an Excel summary."""
    actor_id, context = _resolve_actor_context("export_abstracts_with_pdfs")
    try:
        import zipfile
        import shutil
        from io import BytesIO
        import tempfile
        
        # Get all abstracts with related data
        abstracts = abstract_utils.list_abstracts(
            actor_id=actor_id,
            context=context
        )
        
        # Base path where PDF files are stored
        base_path = current_app.config.get("UPLOAD_FOLDER", "uploads")
        
        # Create a temporary directory to organize files
        temp_dir = tempfile.mkdtemp()
        
        # Organize abstracts by category
        categories = {}
        for abstract in abstracts:
            if abstract.pdf_path and abstract.category:
                category_name = abstract.category.name
                if category_name not in categories:
                    categories[category_name] = []
                categories[category_name].append(abstract)
        
        # Create subdirectories for each category and copy PDF files
        for category_name, abstract_list in categories.items():
            category_dir = os.path.join(temp_dir, category_name)
            os.makedirs(category_dir, exist_ok=True)
            
            for abstract in abstract_list:
                if abstract.pdf_path:
                    # Construct the full path from the base path
                    full_pdf_path = os.path.join(base_path, os.path.basename(abstract.pdf_path))
                    
                    # Verify the file exists before copying
                    if os.path.exists(full_pdf_path):
                        # Get the original filename
                        original_filename = os.path.basename(full_pdf_path)
                        # Create new filename with abstract ID to avoid conflicts
                        new_filename = f"{abstract.id.hex}_{original_filename}"
                        destination_path = os.path.join(category_dir, new_filename)
                        
                        # Copy the PDF file to the category folder
                        shutil.copy2(full_pdf_path, destination_path)
                    else:
                        current_app.logger.warning(f"PDF file not found at path: {full_pdf_path}")
        
        # Generate Excel file with all abstracts data using the existing function
        excel_wb = create_abstracts_excel(abstracts)
        
        # Save the Excel file to the temporary directory root
        excel_path = os.path.join(temp_dir, f"abstracts_summary_{uuid.uuid4().hex[:8]}.xlsx")
        excel_wb.save(excel_path)
        
        # Create a ZIP file from the organized directory structure
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    # Calculate the archive name (path inside the ZIP)
                    archive_name = os.path.relpath(file_path, temp_dir)
                    zip_file.write(file_path, archive_name)
        
        # Clean up the temporary directory
        shutil.rmtree(temp_dir)
        
        # Prepare the ZIP file for download
        zip_buffer.seek(0)
        
        # Log successful export
        log_audit_event(
            event_type="abstract.pdf.excel.zip.export.success",
            user_id=actor_id,
            details={
                "exported_count": len(abstracts),
                "categories_count": len(categories),
                "categories": list(categories.keys())
            },
            ip_address=request.remote_addr
        )
        
        # Send the ZIP file to the user
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=f"abstracts_pdfs_with_excel_{uuid.uuid4().hex[:8]}.zip",
            mimetype='application/zip'
        )
        
    except Exception as exc:
        current_app.logger.exception("Error exporting abstracts with PDFs to ZIP")
        error_msg = f"System error occurred while exporting abstracts with PDFs to ZIP: {str(exc)}"
        log_audit_event(
            event_type="abstract.pdf.excel.zip.export.failed",
            user_id=actor_id,
            details={"error": error_msg, "exception_type": type(exc).__name__, "exception_message": str(exc)},
            ip_address=request.remote_addr
        )
        return jsonify({"error": error_msg}), 400
